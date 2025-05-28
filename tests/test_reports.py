import os
from playwright.sync_api import Page, expect, Route, BrowserContext
from utils.variables import *
from utils.dates import *
from pages.reports import *
from openpyxl import load_workbook
import pytest
import allure
from utils.create_delete_user import create_user, delete_user



@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_check_dates")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_check_dates")
def test_reports_check_dates(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Check begin and end dates in view. today by default"):
        expect(page.locator(FIRST_DATE)).to_have_value(today.strftime("%d/%m/%Y"))
        expect(page.locator(LAST_DATE)).to_have_value(today.strftime("%d/%m/%Y"))

    with allure.step("Switch to yesterday"):
        reports.yesterday.click()

    with allure.step("Check begin and end dates in view"):
        expect(page.locator(FIRST_DATE)).to_have_value(yesterday.strftime("%d/%m/%Y"))
        expect(page.locator(LAST_DATE)).to_have_value(yesterday.strftime("%d/%m/%Y"))

    with allure.step("Click to week"):
        reports.week.click()

    with allure.step("Choose this week"):
        reports.select_period_value("this_week")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_this_week.strftime("%d/%m/%Y"), last_day_this_week.strftime("%d/%m/%Y"))

    with allure.step("Click to week"):
        reports.week.click()

    with allure.step("Choose last week"):
        reports.select_period_value("last_week")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_last_week.strftime("%d/%m/%Y"), last_day_last_week.strftime("%d/%m/%Y"))

    with allure.step("Click to month"):
        reports.month.click()

    with allure.step("Choose this month"):
        reports.select_period_value("this_month")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_this_month.strftime("%d/%m/%Y"), last_day_this_month.strftime("%d/%m/%Y"))

    with allure.step("Click to month"):
        reports.month.click()

    with allure.step("Choose last month"):
        reports.select_period_value("last_month")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_last_month.strftime("%d/%m/%Y"), last_day_last_month.strftime("%d/%m/%Y"))

    with allure.step("Click to quarter"):
        reports.quarter.click()

    with allure.step("Choose this quarter"):
        reports.select_period_value("this_quarter")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_this_quarter.strftime("%d/%m/%Y"), last_day_this_quarter.strftime("%d/%m/%Y"))

    with allure.step("Click to quarter"):
        reports.quarter.click()

    with allure.step("Choose last quarter"):
        reports.select_period_value("last_quarter")

    # with allure.step("Check first and last dates in view."):
    #     communications.assert_check_period_dates(first_day_last_quarter.strftime("%d/%m/%Y"), last_day_last_quarter.strftime("%d/%m/%Y"))

    with allure.step("Click to year"):
        reports.year.click()

    with allure.step("Choose this year"):
        reports.select_period_value("this_year")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_this_year.strftime("%d/%m/%Y"), last_day_this_year.strftime("%d/%m/%Y"))

    with allure.step("Click to year"):
        reports.year.click()

    with allure.step("Choose last year"):
        reports.select_period_value("last_year")

    with allure.step("Check first and last dates in view."):
        reports.assert_check_period_dates(first_day_last_year.strftime("%d/%m/%Y"), last_day_last_year.strftime("%d/%m/%Y"))

    with allure.step("Switch to all time"):
        reports.all_time.click()

    with allure.step("Check begin and end dates is disabled"):
        expect(page.locator(FIRST_DATE)).to_be_disabled()
        expect(page.locator(LAST_DATE)).to_be_disabled()


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_change_dates_in_calendar")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_change_dates_in_calendar. fix https://task.imot.io/browse/DEV-3264")
def test_reports_change_dates_in_calendar(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/02/2022-15/02/2022"):
        reports.choose_period_date("01/02/2022", "15/02/2022")

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report created"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("616")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("2092")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("419")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("3127")

    with allure.step("Choose period 08/02/2022-09/02/2022"):
        reports.choose_period_date("08/02/2022", "09/02/2022")

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report created"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).not_to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("616")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("2092")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("2708")
        expect(page.locator('[data-id="3"]')).not_to_be_visible()

@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_check_calendar_localization")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_check_calendar_localization")
def test_reports_check_calendar_localization(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=False)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with user"):
        reports.auth(LOGIN, PASSWORD)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Click to calendar"):
        page.locator('[date-range="start"]').click()

    with allure.step("Check localization"):
        expect(page.locator('[class="ant-picker-content"]').nth(0)).to_contain_text("ПнВтСрЧтПтСбВс")
        expect(page.locator('[class="ant-picker-content"]').nth(1)).to_contain_text("ПнВтСрЧтПтСбВс")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_save_update_delete_report")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("create report , save , change, save current, delete")
def test_reports_save_update_delete_report(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Press (Save as new)"):
        reports.press_save_as_new()

    with allure.step("Check that opened modal window with empty report name and button (add) disabled"):
        expect(page.locator(INPUT_REPORT_NAME)).to_be_empty()
        expect(page.locator('[class="modal-btns"]').locator(BUTTON_SUBMIT)).to_be_disabled(timeout=wait_until_visible)

    with allure.step("Input report name - auto_test_report"):
        page.locator(INPUT_REPORT_NAME).type("auto_test_report", delay=10)

    with allure.step("Check that button (add) enabled"):
        expect(page.locator('[class="modal-btns"]').locator(BUTTON_SUBMIT)).to_be_enabled(timeout=wait_until_visible)

    with allure.step("Press (save) button"):
        page.locator('[class="modal-btns"]').locator(BUTTON_SUBMIT).click()

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report generated"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("38")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("15")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

        expect(page.locator('[data-testid="templatesReports"]').locator('[class*="body1"]')).to_have_text("auto_test_report")

    with allure.step("Press (save current)"):
        press_save_current(page)

    with allure.step("Check that modal window have report name and (add) button enabled"):
        expect(page.locator(INPUT_REPORT_NAME)).to_have_value("auto_test_report")
        expect(page.locator('[class="modal-btns"]').locator('[type="submit"]')).to_be_enabled()
        expect(page.locator('[class="modal-btns"]').locator('[type="submit"]')).to_have_text("Обновить")

    with allure.step("Press (save) button"):
        page.locator('[class="modal-btns"]').locator('[type="submit"]').click()

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report name not changed"):
        expect(page.locator('[data-testid="templatesReports"]').locator('[class*="body1"]')).to_have_text("auto_test_report")

    with allure.step("Delete report"):
        page.locator('[class=" css-izdlur"]').click()
        page.get_by_text("Удалить шаблон", exact=True).click()

        page.wait_for_selector('[class="modal-btns"]')

        page.get_by_text("Удалить", exact=True).click()
        page.wait_for_timeout(1000)

    with allure.step("Check that report deleted"):
        expect(page.locator('[data-testid="templatesReports"]').locator('[class*="body1"]')).to_have_text("Сохраненные отчеты")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_download_report")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_download_report")
def test_reports_download_report(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report created"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("38")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("15")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Press (Export to Excel) and wait until file will be saved"):
        # Start waiting for the download
        with page.expect_download(timeout=wait_until_visible) as download_info:
            # Perform the action that initiates download
            page.get_by_text("Экспорт в Excel").click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that file was downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        page.wait_for_timeout(500)

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Время"
        assert sheet["A3"].value == "08-02-2022"
        assert sheet["A4"].value == "09-02-2022"
        assert sheet["A5"].value == "10-02-2022"
        assert sheet["A6"].value == "Вошедшие"
        assert sheet["B3"].value == 10
        assert sheet["B4"].value == 38
        assert sheet["B5"].value == 15
        assert sheet["B6"].value == 63
        assert sheet["C3"].value == 10
        assert sheet["C4"].value == 38
        assert sheet["C5"].value == 15
        assert sheet["C6"].value == 63
        assert sheet.max_row == 6
        assert sheet.max_column == 3

    with allure.step("Remove downloaded file"):
        os.remove(path + download.suggested_filename)
        page.wait_for_timeout(500)

    with allure.step("Check that file was removed"):
        assert os.path.isfile(path + download.suggested_filename) == False


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_management_check_search")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_management_check_search")
def test_reports_management_check_search(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Type тест1 in searchstring in menu an press magnifier"):
        page.locator(INPUT_SEARCH).type("тест1", delay=10)
        page.locator(BUTTON_SUBMIT).click()

    with allure.step("Check that search was successful"):
        expect(page.locator(INPUT_SEARCH)).to_have_value("тест1")
        expect(page.get_by_text("тест1")).to_have_count(1)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_management_check")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_management_check")
def test_reports_management_check(base_url, page: Page) -> None:
    period_list = "СегодняВчераТекущая неделяПрошлая неделяТекущий месяцПрошлый месяцТекущий кварталПрошлый кварталТекущий годПрошлый годЗа все времяПроизвольные даты"

    reports = Reports(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with user"):
        reports.auth(LOGIN, PASSWORD)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Go to report management"):
        reports.press_report_management()

    with allure.step("Press (Create report)"):
        reports.press_create_report_in_management()

    with allure.step("Press (save as new)"):
        reports.press_save_as_new()

    with allure.step("Fill report name"):
        page.locator(INPUT_REPORT_NAME).type("auto_test_report", delay=10)

    with allure.step("Press (submit)"):
        page.locator(MODAL_WINDOW).locator(BUTTON_SUBMIT).click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")
        page.wait_for_timeout(500)

    with allure.step("Check alert"):
        reports.check_alert("Отчет сохранен!")

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Go to report management"):
        reports.press_report_management()

    with allure.step("Check table"):
        page.wait_for_timeout(1000)
        expect(page.locator('[role="columnheader"]')).to_have_count(5)
        expect(page.locator('[title="Toggle SortBy"]')).to_have_count(1)
        expect(page.locator('[aria-label="auto_test_report"]')).to_have_count(1)
        expect(page.locator('[aria-label="Изменить название"]')).to_have_count(2)
        expect(page.locator('[class*="styles_buttons"]').locator('[aria-label="Скачать"]')).to_have_count(2)
        expect(page.locator('[aria-label="Создать копию"]')).to_have_count(2)
        expect(page.locator('[aria-label="Удалить"]')).to_have_count(2)
        expect(page.locator('[aria-label="Перейти"]')).to_have_count(2)

    with allure.step("Change report name"):
        page.locator('[aria-label="Изменить название"]').nth(1).click()
        page.locator('[aria-label="Сбросить"]').click()
        page.locator('[aria-label="Изменить название"]').nth(1).click()
        page.locator('[name="reportName"]').clear()
        page.locator('[name="reportName"]').type("changedReportName", delay=10)
        page.locator('[aria-label="Сохранить"]').click()

    with allure.step("Check that report name changed"):
        expect(page.locator('[aria-label="changedReportName"]')).to_have_count(1)

    with allure.step("Click to period menu"):
        page.locator('[class*="styles_periodCell_"]').get_by_text("Сегодня").click()
        page.wait_for_selector(MENU)

    with allure.step("Check content of period options"):
        expect(page.locator('[class*="styles_periodCell_"]').locator(MENU)
               ).to_have_text(period_list)

    with allure.step("Choose arbitrary dates"):
        page.get_by_text("Произвольные даты", exact=True).click()
        page.wait_for_selector('[class="ant-space-item"]')

    with allure.step("Choose dates"):
        reports.choose_period_date(today.strftime("%d/%m/%Y"), today.strftime("%d/%m/%Y"))

    with allure.step("Reload page"):
        reports.reload_page()

    with allure.step("Check that all changes saved"):
        page.wait_for_load_state(state="load", timeout=wait_until_visible)
        expect(page.locator('[aria-label="changedReportName"]')).to_have_count(1)
        expect(page.locator(FIRST_DATE)).to_have_value(today.strftime("%d/%m/%Y"))
        expect(page.locator(LAST_DATE)).to_have_value(today.strftime("%d/%m/%Y"))
#
    with allure.step("Press (Download) and wait until file will be saved"):
        # Start waiting for the download
        with page.expect_download(timeout=wait_until_visible) as download_info:
            # Perform the action that initiates download
            page.locator('[aria-label="Скачать"]').nth(1).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that file was downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 4000 < os.path.getsize(path + download.suggested_filename) < 6500
        page.wait_for_timeout(500)

    with allure.step("Remove downloaded file"):
        os.remove(path + download.suggested_filename)
        page.wait_for_timeout(500)

    with allure.step("Check that file was removed"):
        assert os.path.isfile(path + download.suggested_filename) == False
#
    with allure.step("Make report copy. After create going to main report page"):
        page.locator('[aria-label="Создать копию"]').nth(1).click()
        page.wait_for_selector(MODAL_WINDOW)
        page.locator('[name="newReportName"]').fill("reportCopy")
        page.locator(BUTTON_ACCEPT).click()
        page.wait_for_selector(BUTTON_GENERATE_REPORT)

    with allure.step("Go back to reports"):
        reports.click_reports()

    with allure.step("Go back to report management"):
        reports.press_report_management()

    with allure.step("Check that we have three reports"):
        expect(page.locator('[role="row"]')).to_have_count(4)
        expect(page.locator('[aria-label="changedReportName"]')).to_have_count(1)
        expect(page.locator('[aria-label="reportCopy"]')).to_have_count(1)

    with allure.step("Delete report"):
        page.locator(BUTTON_KORZINA).nth(1).click()
        page.wait_for_selector(BUTTON_ACCEPT)
        page.locator(BUTTON_ACCEPT).click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Check alert"):
        reports.check_alert("Отчет удален!")

    with allure.step("Delete report"):
        page.locator(BUTTON_KORZINA).nth(1).click()
        page.wait_for_selector(BUTTON_ACCEPT)
        page.locator(BUTTON_ACCEPT).click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Check alert"):
        reports.check_alert("Отчет удален!")

    with allure.step("Check that reports deleted"):
        expect(page.locator('[role="row"]')).to_have_count(2)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_report_send_email")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_report_send_email")
def test_report_send_email(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with user"):
        reports.auth(LOGIN, PASSWORD)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Go to report management"):
        reports.press_report_management()

    with allure.step("Press (Send report)"):
        reports.press_send_report()

    with allure.step("Select (Send to email)"):
        reports.choose_where_send_report("Отправлять на почту")

    with allure.step("Check modal window no data text"):
        expect(page.locator('[class*="styles_noDataText"]')).to_contain_text("Этот отчет пока никуда не отправляется")

    with allure.step("Press (Add a new dispatch)"):
        page.locator('[data-testid="addТewNotify_toEmail"]').click()
        page.wait_for_selector('[data-testid="messageBody"]')
        page.wait_for_timeout(500)

    with allure.step("Fill dispatch fields"):
        page.locator('[data-testid="dispatchName_to_toEmail"]').locator('[type="text"]').type("at-email_dispatch", delay=10)
        page.locator('[data-testid="toEmail"]').nth(0).locator('[type="text"]').type("first@mail.com", delay=10)
        page.locator('[data-testid="addEmailBtn"]').click()
        page.locator('[data-testid="toEmail"]').nth(1).locator('[type="text"]').type("second@mail.com", delay=10)
        page.locator('[data-testid="emailSubject_to_toEmail"]').locator('[type="text"]').type("at-email_dispatch", delay=10)
        page.locator('[data-testid="messageBody"]').type("at-email_dispatch", delay=10)
        page.locator('[data-test-id="report_id"]').click()
        page.wait_for_timeout(500)
        page.locator('[aria-label="Текст сообщения"]').click()

    with allure.step("Choose date"):
        # every month
        page.get_by_text("Выберите значение", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждый месяц", exact=True).click()
        page.get_by_text("Выберите день месяца").click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("2", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=20)
        # every week
        page.get_by_text("Каждый месяц", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждую неделю", exact=True).click()
        page.get_by_text("Выберите день недели", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Вторник", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=20)
        # every day
        page.get_by_text("Каждую неделю", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждый день", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=20)

    with allure.step("Press (Send)"):
        page.locator('[data-testid="sendButton_to_toEmail"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Данные сохранены")

    with allure.step("Check modal window no data text"):
        expect(page.locator('[class*="styles_noDataText"]')).to_contain_text("Выберите отправку или добавьте новую")

    with allure.step("Turn of rule"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Вкл/Выкл"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Правило выключено")

    with allure.step("Turn of rule"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Вкл/Выкл"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Правило включено")

    with allure.step("Press (Korzina)"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Удалить"]').click()

    with allure.step("Cancel deleting"):
        page.locator('[class*="styles_buttonsGroup_"]').get_by_text("Отмена")

    with allure.step("Press (Korzina)"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Удалить"]').click()

    with allure.step("Cancel deleting"):
        page.locator('[class*="styles_buttonsGroup_"]').get_by_text("Удалить").click()

    with allure.step("Check alert"):
        reports.check_alert("Правило удалено")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_report_send_telegram")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_report_send_telegram")
def test_report_send_telegram(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    def handle_chat(route: Route):
        json_chat = [{"chatId":-4249734796,"botId":1724205115,"username":"","group":"AT_CHAT"}]
        # fulfill the route with the mock data
        route.fulfill(json=json_chat)
        # Intercept the route
    page.route("**/tg_chats", handle_chat)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with user"):
        reports.auth(LOGIN, PASSWORD)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Go to report management"):
        reports.press_report_management()

    with allure.step("Press (Send report)"):
        reports.press_send_report()

    with allure.step("Select (Send to email)"):
        reports.choose_where_send_report("Отправлять в Telegram")

    with allure.step("Check modal window no data text"):
        expect(page.locator('[class*="styles_noDataText"]')).to_contain_text("Этот отчет пока никуда не отправляется")

    with allure.step("Press (Add a new dispatch)"):
        page.locator('[data-testid="addТewNotify_toTelegram"]').click()
        page.wait_for_selector('[data-testid="messageBody"]')
        page.wait_for_timeout(500)

    with allure.step("Fill dispatch fields"):
        page.locator('[data-testid="dispatchName_to_toTelegram"]').locator('[type="text"]').type("at-telegram_dispatch", delay=10)
        page.locator('[data-testid="selectActiveTgChats"]').locator("svg").click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("AT_CHAT", exact=True).click()
        page.locator('[data-testid="messageBody"]').type("at-telegram_dispatch", delay=10)
        page.locator('[data-test-id="report_id"]').click()
        page.wait_for_timeout(500)

    with allure.step("Choose date"):
        # every month
        page.get_by_text("Выберите значение", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждый месяц", exact=True).click()
        page.get_by_text("Выберите день месяца").click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("2", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=10)
        # every week
        page.get_by_text("Каждый месяц", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждую неделю", exact=True).click()
        page.get_by_text("Выберите день недели", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Вторник", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=10)
        # every day
        page.get_by_text("Каждую неделю", exact=True).click()
        page.wait_for_selector(MENU)
        page.locator(MENU).get_by_text("Каждый день", exact=True).click()
        page.locator('[placeholder="00:00"]').type("1111", delay=10)

    with allure.step("Press (Send)"):
        page.locator('[data-testid="sendButton_to_toTelegram"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Данные сохранены")

    with allure.step("Check modal window no data text"):
        expect(page.locator('[class*="styles_noDataText"]')).to_contain_text("Выберите отправку или добавьте новую")

    with allure.step("Turn of rule"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Вкл/Выкл"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Правило выключено")

    with allure.step("Turn of rule"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Вкл/Выкл"]').click()

    with allure.step("Check alert"):
        reports.check_alert("Правило включено")

    with allure.step("Press (Korzina)"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Удалить"]').click()

    with allure.step("Cancel deleting"):
        page.locator('[class*="styles_buttonsGroup_"]').get_by_text("Отмена")

    with allure.step("Press (Korzina)"):
        page.locator(MODAL_WINDOW).locator('[aria-label="Удалить"]').click()

    with allure.step("Cancel deleting"):
        page.locator('[class*="styles_buttonsGroup_"]').get_by_text("Удалить").click()

    with allure.step("Check alert"):
        reports.check_alert("Правило удалено")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_go_to_communications_from_report")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_go_to_communications_from_report")
def test_reports_go_to_communications_from_report(base_url, page: Page, context: BrowserContext) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Go to calls from report"):
        with context.expect_page() as new_tab_event:
            page.locator('[title="3130"]').first.click()
            new_tab = new_tab_event.value

    with allure.step("Check"):
        expect(new_tab.locator(INPUT_BY_TAGS)).to_be_visible(timeout=wait_until_visible)
        expect(new_tab.locator(BUTTON_FIND_COMMUNICATIONS)).to_be_visible(timeout=wait_until_visible)
        expect(new_tab.locator(BUTTON_CLEAR)).to_be_visible(timeout=wait_until_visible)
        expect(new_tab.locator(BUTTON_CALLS_LIST_DOWNLOAD)).to_have_count(2, timeout=wait_until_visible)
        expect(new_tab.locator(BUTTON_RETAG)).to_have_count(2, timeout=wait_until_visible)
        expect(new_tab.locator(BUTTON_CALLS_ACTION)).to_have_count(2, timeout=wait_until_visible)
        expect(new_tab.locator(NAYDENO_ZVONKOV).first).to_have_text("Найдено коммуникаций 3130 из 3130")

    with allure.step("Close context"):
        new_tab.close()


# rows
@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_1_without_grouping")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_1_without_grouping")
def test_reports_row_1_without_grouping(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Make row without grouping"):
        reports.choose_grouping_without_parameters("row", "1", "Без группировки")

    #fill_column_by_communication("0", page)

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report generated"):
        expect(page.locator('[aria-label="Без группировки"]')).to_have_count(2)

        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand report parameters"):
        reports.expand_report()

    with allure.step("Check checklist exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

    with allure.step("Check row is without grouping"):
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Без группировки")

    with allure.step("Check column is by communications"):
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_1_time_4_values")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_1_time_4_values")
def test_reports_row_1_time_4_values(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Choose first option (by days)"):
        reports.choose_row_by_date("1", "Времени", "По дням")
        #fill_row_by_date("1", "Времени", "По дням", page)

    with allure.step("Fill column (by communications) by default"):
        #fill_column_by_communication("0", page)
        reports.choose_grouping_without_parameters("column","0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("38")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("15")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Времени")
        expect(page.locator('[data-testid="report_rows_row_1_time"]')).to_have_text("По дням")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")

    with allure.step("Choose second option (by weeks)"):
        reports.choose_row_by_date("1", "Времени", "По неделям")
        #fill_row_by_date("1", "Времени", "По неделям", page)

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="6 неделя, 2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Времени")
        expect(page.locator('[data-testid="report_rows_row_1_time"]')).to_have_text("По неделям")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")

    with allure.step("Choose second option (by month)"):
        reports.choose_row_by_date("1", "Времени", "По месяцам")
        #fill_row_by_date("1", "Времени", "По месяцам", page)

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Времени")
        expect(page.locator('[data-testid="report_rows_row_1_time"]')).to_have_text("По месяцам")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")

    with allure.step("Choose second option (by hours)"):
        reports.choose_row_by_date("1", "Времени", "По часам")
        #fill_row_by_date("1", "Времени", "По часам", page)

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="08-02-2022 10:00 - 11:00"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("4")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("3")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("2")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("1")
        expect(page.locator('[data-id="4"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("3")

    with allure.step("Scroll down"):
        page.locator('[class*="MuiDataGrid-scrollbar--vertical"]').click()
        page.mouse.wheel(delta_x=0, delta_y=10000)

    with allure.step("Check that report generated and all values okey (last)"):
        expect(page.locator('[data-id="17"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("7")
        expect(page.locator('[data-id="18"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Времени")
        expect(page.locator('[data-testid="report_rows_row_1_time"]')).to_have_text("По часам")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_1_communications")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_1_communications")
def test_reports_row_1_communications(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Choose row by communications"):
        reports.choose_grouping_without_parameters("row","1", "Коммуникации")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Scroll down"):
        page.locator('[class*="MuiDataGrid-scrollbar--vertical"]').click()
        page.mouse.wheel(delta_x=0, delta_y=10000)

    with allure.step("Check that report generated and all values okey (last)"):
        expect(page.locator('[data-id="63"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Коммуникации")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_1_operator_phone")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_1_operator_phone")
def test_reports_row_1_operator_phone(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Fill row by operators phone"):
        reports.choose_grouping_without_parameters("row", "1", "Номеру сотрудника")

    with allure.step("Fill column by communication"):
        #fill_column_by_communication("0", page)
        reports.choose_grouping_without_parameters("column", "0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="251"]')).to_be_visible()
        expect(page.locator('[aria-label="266"]')).to_be_visible()
        expect(page.locator('[aria-label="122"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("5")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("1")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("4")

    with allure.step("Scroll down"):
        page.locator('[class*="MuiDataGrid-scrollbar--vertical"]').click()
        page.mouse.wheel(delta_x=0, delta_y=10000)

    with allure.step("Check that report generated and all values okey (last)"):
        expect(page.locator('[data-id="20"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("1")
        expect(page.locator('[data-id="21"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Номеру сотрудника")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_1_client_phone")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_1_client_phone")
def test_reports_row_1_client_phone(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Fill row by client number"):
        reports.choose_grouping_without_parameters("row","1", "Номеру клиента")

    with allure.step("Fill column by communication"):
        reports.choose_grouping_without_parameters("column", "0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="89164476315"]')).to_be_visible()
        expect(page.locator('[aria-label="89645653870"]')).to_be_visible()
        expect(page.locator('[aria-label="89637275939"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("6")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("4")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("2")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("6")

    with allure.step("Scroll down"):
        page.locator('[class*="MuiDataGrid-scrollbar--vertical"]').click()
        page.mouse.wheel(delta_x=0, delta_y=10000)

    with allure.step("Check that report generated and all values okey (last)"):
        expect(page.locator('[data-id="31"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("1")
        expect(page.locator('[data-id="32"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()
        # check row
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Номеру клиента")
        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_2_tag_list")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_2_tag_list")
def test_reports_row_2_tag_list(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Add first row by tag list"):
        reports.choose_tag_list("row", "1", "direction")
        #fill_row_by_tag_list("1", "По списку тегов", "direction", page)

    with allure.step("Press (Add row)"):
        reports.press_add_row()

    with allure.step("Add second row by tag list"):
        reports.choose_tag_list("row", "2", "hangup")
        #fill_row_by_tag_list("2", "По списку тегов", "hangup", page)

    with allure.step("Uncheck check box in 2nd row"):
        reports.click_checkbox_in_tag_list("row", "2")

        expect(page.locator('[data-testid="report_rows_row_2_tagListCheckbox"]')).not_to_be_checked()

    with allure.step("Fill column by communications"):
        reports.choose_grouping_without_parameters("column", "0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        #expect(page.locator('[aria-label="direction + hangup, CALLID"]')).to_be_visible()
        expect(page.locator('[aria-label="direction + hangup"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("По списку тегов")
        expect(page.locator('[data-testid="report_rows_row_2_select"]')).to_have_text("По списку тегов")

        # check tagSelect
        expect(page.locator('[data-testid="report_rows_row_1__tagListValues"]')).to_have_text("direction")
        #expect(page.locator('[data-testid="report_rows_row_2__tagListValues"]')).to_have_text("hangupCALLID")
        expect(page.locator('[data-testid="report_rows_row_2__tagListValues"]')).to_have_text("hangup")
        # check checkboxes
        expect(page.locator('[data-testid="report_rows_row_1_tagListCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_rows_row_2_tagListCheckbox"]')).not_to_be_checked()


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_row_4_tag_and_value")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_row_4_tag_and_value")
def test_reports_row_4_tag_and_value(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Чек лист звонка (не трогать, автотест повзязан на баллы которые получаются в результате применения этого чек листа)"):
        reports.add_checklist_to_report("Чек лист звонка (не трогать, автотест повзязан на баллы которые получаются в результате применения этого чек листа)")

    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("row", "1", "Тегу и значениям", "direction", "incoming")
        #fill_row_by_tag_and_value("1", "Тегу и значениям", "direction", "incoming", page)

    with allure.step("Press (Add row)"):
        reports.press_add_row()

    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("row", "2", "Тегу и значениям", "hangup", "operator")
        #fill_row_by_tag_and_value("2", "Тегу и значениям", "hangup", "operator", page)

    with allure.step("Press (Add row)"):
        reports.press_add_row()

    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("row", "3", "Тегу и значениям", "CALLID", "Выбрать все")
        #fill_row_by_tag_and_value("3", "Тегу и значениям", "CALLID", "Выбрать все", page)

    with allure.step("Uncheck checkbox in 3d row"):
        reports.click_checkbox_in_tag_and_value("row", "3")

        expect(page.locator('[data-testid="report_rows_row_3_tagCheckbox"]')).not_to_be_checked()

    with allure.step("Press (Add row)"):
        reports.press_add_row()

    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("row", "4", "Тегу и значениям", "queue", "Выбрать все")
        #fill_row_by_tag_and_value("4", "Тегу и значениям", "queue", "Выбрать все", page)

    with allure.step("Uncheck checkbox in 4th row"):
        reports.click_checkbox_in_tag_and_value("row", "4")

        expect(page.locator('[data-testid="report_rows_row_4_tagCheckbox"]')).not_to_be_checked()

    with allure.step("Fill column with communications"):
        #fill_column_by_communication("0", page)
        reports.choose_grouping_without_parameters("column", "0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="incoming + operator + CALLID + queue"]')).to_be_visible()

        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("193")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("193")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Чек лист звонка (не трогать, автотест повзязан на баллы которые получаются в результате применения этого чек листа)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_rows_row_1_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_rows_row_2_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_rows_row_3_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_rows_row_4_select"]')).to_have_text("Тегу и значениям")
        # check tagSelect
        expect(page.locator('[data-testid="report_rows_row_1_tagSelect"]')).to_have_text("direction")
        expect(page.locator('[data-testid="report_rows_row_2_tagSelect"]')).to_have_text("hangup")
        expect(page.locator('[data-testid="report_rows_row_3_tagSelect"]')).to_have_text("CALLID")
        expect(page.locator('[data-testid="report_rows_row_4_tagSelect"]')).to_have_text("queue")
        # check tagValues
        expect(page.locator('[data-testid="report_rows_row_1_tagValues"]')).to_have_text("incoming")
        expect(page.locator('[data-testid="report_rows_row_2_tagValues"]')).to_have_text("operator")
        expect(page.locator('[data-testid="report_rows_row_3_tagValues"]')).to_have_text("Выбрать все")
        expect(page.locator('[data-testid="report_rows_row_4_tagValues"]')).to_have_text("Выбрать все")
        # check checkboxes
        expect(page.locator('[data-testid="report_rows_row_1_tagCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_rows_row_2_tagCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_rows_row_3_tagCheckbox"]')).not_to_be_checked()
        expect(page.locator('[data-testid="report_rows_row_4_tagCheckbox"]')).not_to_be_checked()


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_column_1_communications")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_column_1_communications")
def test_reports_column_1_communications(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Add column with communications"):
        #fill_column_by_communication("0",page)
        reports.choose_grouping_without_parameters("column", "0", "По количеству коммуникаций")

    with allure.step("Press (Generate report)"):
        reports.press_generate_report()

    with allure.step("Check that report generated and all values okey"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("38")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("15")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По количеству коммуникаций")


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_column_4_filter")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_column_4_filter")
def test_reports_column_4_filter(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    # 0
    with allure.step("Fill 0 column with filter"):
        #reports.fill_column_by_exact_filter("0", "zero", "Адрес подключения", "девятнадцать")
        reports.fill_column_by_exact_filter("0", "zero", "Отдел", "Контроля Качества")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 1
    with allure.step("Fill 1 column fith filter"):
        reports.fill_column_by_exact_filter("1", "first", "Клиент-должность", "Монтажник Восток")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 2
    with allure.step("Fill 2 column fith filter"):
        reports.fill_column_by_exact_filter("2", "second", "Должность", "Бухгалтер")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 3
    with allure.step("Fill 3 column fith filter"):
        reports.fill_column_by_exact_filter("3", "third", "Клиент", "Александр")

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report generated"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("6") #5
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10") #8
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("6") #6
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("22") #19

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("Точный фильтр")
        expect(page.locator('[data-testid="report_columns_column_1_select"]')).to_have_text("Точный фильтр")
        expect(page.locator('[data-testid="report_columns_column_2_select"]')).to_have_text("Точный фильтр")
        expect(page.locator('[data-testid="report_columns_column_3_select"]')).to_have_text("Точный фильтр")
        # check tagSelect
        expect(page.locator('[data-testid="report_columns_column_0_searchInput"]').locator('[type="text"]')).to_have_value("zero")
        expect(page.locator('[data-testid="report_columns_column_1_searchInput"]').locator('[type="text"]')).to_have_value("first")
        expect(page.locator('[data-testid="report_columns_column_2_searchInput"]').locator('[type="text"]')).to_have_value("second")
        expect(page.locator('[data-testid="report_columns_column_3_searchInput"]').locator('[type="text"]')).to_have_value("third")

        # check tags
        expect(page.locator('[data-testid="report_columns"]').get_by_text("Контроля Качества")).to_be_visible()
        expect(page.locator('[data-testid="report_columns"]').get_by_text("Монтажник Восток")).to_be_visible()
        expect(page.locator('[data-testid="report_columns"]').get_by_text("Бухгалтер")).to_be_visible()
        expect(page.locator('[data-testid="report_columns"]').get_by_text("Александр")).to_be_visible()


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_column_4_tag_and_value")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_column_4_tag_and_value")
def test_reports_column_4_tag_and_value(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    # 0
    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("column", "0", "Тегу и значениям", "direction", "outgoing")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 1
    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("column", "1", "Тегу и значениям", "hangup", "operator")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 2
    with allure.step("Add row with Tag and value"):
        reports.choose_tag_and_value("column", "2", "Тегу и значениям", "CALLID", "Выбрать все")

    with allure.step("Uncheck checkbox in 2rd row"):
        reports.click_checkbox_in_tag_and_value("column","2")

        expect(page.locator('[data-testid="report_columns_column_2_tagCheckbox"]')).not_to_be_checked()

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 3
    with allure.step("Add column with Tag and value"):
        reports.choose_tag_and_value("column", "3", "Тегу и значениям", "queue", "Выбрать все")

    with allure.step("Uncheck checkbox in 3rd column"):
        reports.click_checkbox_in_tag_and_value("column","3")

        expect(page.locator('[data-testid="report_columns_column_3_tagCheckbox"]')).not_to_be_checked()

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report generated"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("9")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("36")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("13")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("58")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_columns_column_1_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_columns_column_2_select"]')).to_have_text("Тегу и значениям")
        expect(page.locator('[data-testid="report_columns_column_3_select"]')).to_have_text("Тегу и значениям")
        # check tagSelect
        expect(page.locator('[data-testid="report_columns_column_0_tagSelect"]')).to_have_text("direction")
        expect(page.locator('[data-testid="report_columns_column_1_tagSelect"]')).to_have_text("hangup")
        expect(page.locator('[data-testid="report_columns_column_2_tagSelect"]')).to_have_text("CALLID")
        expect(page.locator('[data-testid="report_columns_column_3_tagSelect"]')).to_have_text("queue")
        # check tagValues
        expect(page.locator('[data-testid="report_columns_column_0_tagValues"]')).to_have_text("outgoing")
        expect(page.locator('[data-testid="report_columns_column_1_tagValues"]')).to_have_text("operator")
        expect(page.locator('[data-testid="report_columns_column_2_tagValues"]')).to_have_text("Выбрать все")
        expect(page.locator('[data-testid="report_columns_column_3_tagValues"]')).to_have_text("Выбрать все")
        # check checkboxes
        expect(page.locator('[data-testid="report_columns_column_0_tagCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_1_tagCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_2_tagCheckbox"]')).not_to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_3_tagCheckbox"]')).not_to_be_checked()


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_column_4_tag_list")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_column_4_tag_list")
def test_reports_column_4_tag_list(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    with allure.step("Add filter check-list : Второй чеклист (тоже нужен для автотестов, не трогать)"):
        reports.add_checklist_to_report("Второй чеклист (тоже нужен для автотестов, не трогать)")

    # 0
    with allure.step("Add first row by tag list"):
        reports.choose_tag_list("column", "0", "11", "asterisk_context")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 1
    with allure.step("Add column with Tag list"):
        reports.choose_tag_list("column", "1", "CALLID", "direction")

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 2
    with allure.step("Add column with Tag list"):
        reports.choose_tag_list("column", "2", "hangup", "k")

    with allure.step("Uncheck checkbox in 2nd column"):
        reports.click_checkbox_in_tag_list("column", "2")

        expect(page.locator('[data-testid="report_columns_column_2_tagListCheckbox"]')).not_to_be_checked()

    with allure.step("Press (Add column)"):
        reports.press_add_column()

    # 3
    with allure.step("Add column with Tag list"):
        reports.choose_tag_list("column", "3", "multi value", "multi value number")

    with allure.step("Uncheck checkbox in 3rd column"):
        reports.click_checkbox_in_tag_list("column", "3")

        expect(page.locator('[data-testid="report_columns_column_3_tagListCheckbox"]')).not_to_be_checked()

    with allure.step("Press generate report"):
        reports.press_generate_report()

    with allure.step("Check that report generated"):
        expect(page.locator('[aria-label="08-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="09-02-2022"]')).to_be_visible()
        expect(page.locator('[aria-label="10-02-2022"]')).to_be_visible()
        expect(page.locator('[data-id="0"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("10")
        expect(page.locator('[data-id="1"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("38")
        expect(page.locator('[data-id="2"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("15")
        expect(page.locator('[data-id="3"]').locator('[data-field="row_sum_calls_count"]')).to_have_text("63")

    with allure.step("Expand reports parameters"):
        reports.expand_report()

    with allure.step("Check that all parameters exists"):
        expect(page.locator('[aria-label="Remove Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_be_visible()

        # check column
        expect(page.locator('[data-testid="report_columns_column_0_select"]')).to_have_text("По списку тегов")
        expect(page.locator('[data-testid="report_columns_column_1_select"]')).to_have_text("По списку тегов")
        expect(page.locator('[data-testid="report_columns_column_2_select"]')).to_have_text("По списку тегов")
        expect(page.locator('[data-testid="report_columns_column_3_select"]')).to_have_text("По списку тегов")
        # check tagSelect
        expect(page.locator('[data-testid="report_columns_column_0__tagListValues"]')).to_have_text("11asterisk_context")
        expect(page.locator('[data-testid="report_columns_column_1__tagListValues"]')).to_have_text("CALLIDdirection")
        expect(page.locator('[data-testid="report_columns_column_2__tagListValues"]')).to_have_text("hangupk")
        expect(page.locator('[data-testid="report_columns_column_3__tagListValues"]')).to_have_text("multi valuemulti value number")

        # check checkboxes
        expect(page.locator('[data-testid="report_columns_column_0_tagListCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_1_tagListCheckbox"]')).to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_2_tagListCheckbox"]')).not_to_be_checked()
        expect(page.locator('[data-testid="report_columns_column_3_tagListCheckbox"]')).not_to_be_checked()

        expect(page.locator('[data-testid="report_columns_column_2_tagListInput"]').locator('[type="text"]')).to_have_value("hangup, k")
        expect(page.locator('[data-testid="report_columns_column_3_tagListInput"]').locator('[type="text"]')).to_have_value("multi value, multi value number")


# additional params

@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_content")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_content")
def test_reports_additional_params_content(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with admin"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Check content of modal window for rows. 21 checkbox, 7 add params, 2 buttons"):
        expect(page.locator(MODAL_WINDOW).locator('[type="checkbox"]')).to_have_count(29)
        expect(page.locator(BUTTON_TAG_VALUE_IN_ADDITIONAL_PARAMS)).to_have_count(1)
        expect(page.locator('[data-testid="checklistChange"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistChangePercent"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistQuestionChange"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistQuestionChangePercent"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistAnswerAvg"]')).to_have_count(1)
        expect(page.locator('[data-testid="commentary"]')).to_have_count(1)
        expect(page.get_by_role("button", name="Применить")).to_have_count(1)
        expect(page.get_by_role("button", name="Применить")).to_have_attribute("tabindex", "0")
        expect(page.get_by_role("button", name="Отмена")).to_have_count(1)
        expect(page.get_by_role("button", name="Отмена")).to_have_attribute("tabindex", "0")

    with allure.step("Close modal window"):
        page.locator(BUTTON_CROSS).click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Click additional params for columns"):
        reports.click_gear_in_columns("0")

    with allure.step("Check content of modal window for columns. 22 checkbox, 8 add params, 2 buttons"):
        expect(page.locator(MODAL_WINDOW).locator('[type="checkbox"]')).to_have_count(31)
        expect(page.locator(BUTTON_TAG_VALUE_IN_ADDITIONAL_PARAMS)).to_have_count(1)
        expect(page.locator('[data-testid="checklistChange"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistChangePercent"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistQuestionChange"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistQuestionChangePercent"]')).to_have_count(1)
        expect(page.locator('[data-testid="checklistAnswerAvg"]')).to_have_count(1)
        expect(page.locator('[data-testid="commentary"]')).to_have_count(1)
        expect(page.locator('[data-testid="percentSource"]')).to_have_count(1)
        expect(page.get_by_role("button", name="Применить")).to_have_count(1)
        expect(page.get_by_role("button", name="Применить")).to_have_attribute("tabindex", "0")
        expect(page.get_by_role("button", name="Отмена")).to_have_count(1)
        expect(page.get_by_role("button", name="Отмена")).to_have_attribute("tabindex", "0")

    with allure.step("Close modal window"):
        page.locator(BUTTON_CROSS).click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)

# additional params combination

@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_tag_value")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_tag_value. For rows and columns")
def test_reports_additional_params_tag_value(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("asterisk_context")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("asterisk_context")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_asterisk_context_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="asterisk_context"]')).to_have_count(2)
        # check that contain text
        expect(page.get_by_text("ecotelecom-support")).to_have_count(8)

#
@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_avg_number_tag_value")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_avg_number_tag_value. For rows and columns")
def test_reports_additional_params_avg_number_tag_value(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_avg_number_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_avg_number_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("multi value number")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_avg_number_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_avg_number_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("multi value number")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[title="5061.89"]')).to_have_count(2)
        expect(page.locator('[title="5092.72"]')).to_have_count(2)
        expect(page.locator('[title="4956.51"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="multi value number"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_sum_number_tag_value")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_sum_number_tag_value. For rows and columns")
def test_reports_additional_params_sum_number_tag_value(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_sum_number_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_sum_number_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("multi value number")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_sum_number_tag_value()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click tag value in additional params and wait for select"):
        reports.click_add_param_sum_number_tag_value()

    with allure.step("Choose tag in select"):
        reports.type_value_to_select("multi value number")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[title="0"]')).to_have_count(4)
        expect(page.locator('[title="7292773"]')).to_have_count(2)
        expect(page.locator('[title="2076777"]')).to_have_count(2)
        expect(page.locator('[title="9369550"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="multi value number"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checklist_point")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checklist_point. For rows and columns")
def test_reports_additional_params_checklist_point(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist point in additional params and wait for select"):
        reports.click_add_param_checklist_point()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist point in additional params and wait for select"):
        reports.click_add_param_checklist_point()

    with allure.step("Choose checklist in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist point in additional params and wait for select"):
        reports.click_add_param_checklist_point()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist point in additional params and wait for select"):
        reports.click_add_param_checklist_point()

    with allure.step("Choose checklist in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_Второй чеклист (тоже нужен для автотестов, не трогать)_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_have_count(2)
        # check avarage checklist value (0, 1.05, 0, -)
        expect(page.locator('[title="0.63"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checklist_point_percent")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checklist_point_percent. For rows and columns")
def test_reports_additional_params_checklist_point_percent(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist point percent in additional params and wait for select"):
        reports.click_add_param_checklist_point_percent()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist point percent in additional params and wait for select"):
        reports.click_add_param_checklist_point_percent()

    with allure.step("Choose checklist in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist point percent in additional params and wait for select"):
        reports.click_add_param_checklist_point_percent()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist point percent in additional params and wait for select"):
        reports.click_add_param_checklist_point_percent()

    with allure.step("Choose checklist in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_Второй чеклист (тоже нужен для автотестов, не трогать)_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Второй чеклист (тоже нужен для автотестов, не трогать)"]')).to_have_count(2)
        # check avarage checklist value (0 %, 10.53 %, 0 %, -)
        expect(page.locator('[title="6.35 %"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checklist_question_point")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checklist_question_point. For rows and columns")
def test_reports_additional_params_checklist_question_point(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist question point in additional params and wait for select"):
        reports.click_add_param_checklist_question_point()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point in additional params and wait for select"):
        reports.click_add_param_checklist_question_point()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_question_point()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point in additional params and wait for select"):
        reports.click_add_param_checklist_question_point()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?"]')).to_have_count(2)
        # check avarage checklist value (0, 1.05, 0, -)
        expect(page.locator('[title="0.63"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checklist_question_point_in_percent")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checklist_question_point_in_percent. For rows and columns")
def test_reports_additional_params_checklist_question_point_in_percent(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist question point in additional params and wait for select"):
        reports.click_add_param_checklist_question_point_percent()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_question_point_percent()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_question_point_percent()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_question_point_percent()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?"]')).to_have_count(2)
        # check avarage checklist value (0 %, 10.53 %, 0 %, -)
        expect(page.locator('[title="6.35 %"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checklist_frequent_answer_for_question")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checklist_frequent_answer_for_question. For rows and columns")
def test_reports_additional_params_checklist_frequent_answer_for_question(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist frequent answer in additional params and wait for select"):
        reports.click_add_param_checklist_frequent_answer_for_question()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_frequent_answer_for_question()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_checklist_frequent_answer_for_question()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist frequent answer in additional params and wait for select"):
        reports.click_add_param_checklist_frequent_answer_for_question()

    with allure.step("Choose checklist question in select"):
        reports.type_value_to_select("Второй чеклист (тоже нужен для автотестов, не трогать)")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="Второй чеклист (тоже нужен для автотестов, не трогать) / Мат?"]')).to_have_count(2)
        # check
        expect(page.locator('[title="да"]')).to_have_count(4)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_comment")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_comment. For rows and columns")
def test_reports_additional_params_comment(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Click checklist question point in additional params and wait for select"):
        reports.click_add_param_comment()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_comment()

    with allure.step("Choose comment in select"):
        reports.type_value_to_select("2222")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_comment()

    with allure.step("Delete select by using basket button and wait until deleted"):
        reports.delete_select()

    with allure.step("Click checklist question point  in additional params and wait for select"):
        reports.click_add_param_comment()

    with allure.step("Choose comment in select"):
        reports.type_value_to_select("2222")

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        expect(page.locator('[data-field="calls_count_2222_0_0"]')).to_have_count(6)
        # check headers
        expect(page.locator('[aria-label="Коммуникации"]')).to_have_count(2)
        # check headers
        expect(page.locator('[aria-label="2222"]')).to_have_count(2)
        #expect(page.locator('[data-id="4"]').locator('[title="qqqq,3333"]')).to_have_count(1)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_percentage")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_percentage. 5 checkboxes with percentage. For rows and columns")
def test_reports_additional_params_checkboxes_percentage(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 5 percentage checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_PERCENTAGE_FROM_REPORT_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_ROW_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_COLUMN_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_ROW_CELL_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_COLUMN_CELL_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 5 percentage checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_PERCENTAGE_FROM_REPORT_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_ROW_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_COLUMN_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_ROW_CELL_ADD_PARAMS).check()
        page.locator(CHECKBOX_PERCENTAGE_FROM_COLUMN_CELL_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        # check headers
        expect(page.locator('[aria-label="Доля (%) от коммуникаций, вошедших в отчет"]')).to_have_count(2)
        expect(page.locator('[aria-label="Доля (%) от коммуникаций, вошедших в строку"]')).to_have_count(2)
        expect(page.locator('[aria-label="Доля (%) от коммуникаций, вошедших в колонку"]')).to_have_count(2)
        expect(page.locator('[aria-label="Доля (%) от коммуникаций, вошедших в ячейку строки"]')).to_have_count(2)
        expect(page.locator('[aria-label="Доля (%) от коммуникаций, вошедших в ячейку колонки"]')).to_have_count(2)
        # check sum
        expect(page.locator('[title="19.68 %"]')).to_have_count(4)
        expect(page.locator('[title="66.84 %"]')).to_have_count(4)
        expect(page.locator('[title="13.39 %"]')).to_have_count(4)
        expect(page.locator('[title="0.1 %"]')).to_have_count(4)
        expect(page.locator('[title="100 %"]')).to_have_count(24)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_points")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_points. 4 checkboxes with points. For rows and columns")
def test_reports_additional_params_checkboxes_points(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 4 points checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_AVERAGE_POINT_CHECKLIST_ADD_PARAMS).check()
        page.locator(CHECKBOX_AVERAGE_POINT_CHECKLIST_PERCENT_ADD_PARAMS).check()
        page.locator(CHECKBOX_POINTS_SUM_ADD_PARAMS).check()
        page.locator(CHECKBOX_POINTS_MAX_SUM_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 4 points checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_AVERAGE_POINT_CHECKLIST_ADD_PARAMS).check()
        page.locator(CHECKBOX_AVERAGE_POINT_CHECKLIST_PERCENT_ADD_PARAMS).check()
        page.locator(CHECKBOX_POINTS_SUM_ADD_PARAMS).check()
        page.locator(CHECKBOX_POINTS_MAX_SUM_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        # check headers
        expect(page.locator('[aria-label="Сумма баллов по чек-листам"]')).to_have_count(2)
        expect(page.locator('[aria-label="Сумма максимальных баллов по чек-листам"]')).to_have_count(2)
        expect(page.locator('[aria-label="Средний балл"]')).to_have_count(2)
        expect(page.locator('[aria-label="Средний балл в %"]')).to_have_count(2)
        # check sum
        expect(page.locator('[title="14019"]')).to_have_count(2)
        expect(page.locator('[title="49481"]')).to_have_count(2)
        expect(page.locator('[title="10.99"]')).to_have_count(2)
        expect(page.locator('[title="28.33 %"]')).to_have_count(2)
        #expect(page.locator('[title="29.33"]')).to_have_count(1) #this is bug


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_talk_time")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_talk_time. 4 checkboxes with talk time. For rows and columns")
def test_reports_additional_params_checkboxes_talk_time(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 4 percentage checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_CLIENT_TALK_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_CLIENT_TALK_TIME_PERCENT_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATOR_TALK_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATOR_TALK_TIME_PERCENT_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 4 percentage checkboxes"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_CLIENT_TALK_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_CLIENT_TALK_TIME_PERCENT_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATOR_TALK_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATOR_TALK_TIME_PERCENT_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        # check headers
        expect(page.locator('[aria-label="Время разговора оператора"]')).to_have_count(2)
        expect(page.locator('[aria-label="Время разговора клиента"]')).to_have_count(2)
        expect(page.locator('[aria-label="% разговора оператора"]')).to_have_count(2)
        expect(page.locator('[aria-label="% разговора клиента"]')).to_have_count(2)
        # check sum
        page.wait_for_timeout(15000)
        expect(page.locator('[title="28:33:47"]')).to_have_count(2)
        expect(page.locator('[title="27:44:26"]')).to_have_count(2)
        expect(page.locator('[title="35.34 %"]')).to_have_count(2)
        expect(page.locator('[title="34.33 %"]')).to_have_count(2)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_silence_duration")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_silence_duration. 2 checkboxes with silence duration. For rows and columns")
def test_reports_additional_params_checkboxes_silence_duration(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 2 silence duration"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_SILENCE_DURATION_ADD_PARAMS).check()
        page.locator(CHECKBOX_SILENCE_DURATION_PERCENT_ADD_PARAMS).check()


    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 2 silence duration"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_SILENCE_DURATION_ADD_PARAMS).check()
        page.locator(CHECKBOX_SILENCE_DURATION_PERCENT_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        # check headers
        expect(page.locator('[aria-label="Продолжительность тишины"]')).to_have_count(2)
        expect(page.locator('[aria-label="% продолжительности тишины"]')).to_have_count(2)
        # check sum
        expect(page.locator('[title="24:30:36"]')).to_have_count(2)
        expect(page.locator('[title="30.33 %"]')).to_have_count(2)

@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_phones")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_phones. 2 checkboxes with phones. For rows and columns")
def test_reports_additional_params_checkboxes_phones(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 2 phones"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_CLIENTS_PHONES_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATORS_PHONES_ADD_PARAMS).check()


    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 2 phones"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_CLIENTS_PHONES_ADD_PARAMS).check()
        page.locator(CHECKBOX_OPERATORS_PHONES_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):
        # check headers
        expect(page.locator('[aria-label="Телефоны клиентов"]')).to_have_count(2)
        expect(page.locator('[aria-label="Телефоны сотрудников"]')).to_have_count(2)
        # check sum
        expect(page.get_by_text("89672838036")).to_have_count(4)
        expect(page.get_by_text("4958017857")).to_have_count(6)


@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_additional_params_checkboxes_sum_time_first_last_time")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_additional_params_checkboxes_sum_time_first_last_time. 2 checkboxes with sum time and first last time. For rows and columns")
def test_reports_additional_params_checkboxes_sum_time_first_last_time(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 01/01/2022-31/12/2022"):
        reports.choose_period_date("01/01/2022", "31/12/2022")

    # for row
    with allure.step("Click additional params for rows"):
        reports.click_gear_in_rows()

    with allure.step("Unclick communications checkbox and click 2 sum time and first/last time"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_SUM_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_FIRST_COMM_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_LAST_COMM_TIME_ADD_PARAMS).check()


    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # for column

    with allure.step("Click additional params for column"):
        reports.click_gear_in_columns("0")

    with allure.step("Unclick communications checkbox and click 2 sum time and first/last time"):
        page.locator(CHECKBOX_COMMUNICATIONS_ADD_PARAMS).uncheck()
        page.locator(CHECKBOX_SUM_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_FIRST_COMM_TIME_ADD_PARAMS).check()
        page.locator(CHECKBOX_LAST_COMM_TIME_ADD_PARAMS).check()

    with allure.step("Click (Apply)"):
        reports.click_apply_in_additional_params()

    # generate report
    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("check"):

        #  REMINDER: add Средняя длительность коммуникаций!!! https://task.imot.io/browse/DEV-3365

        # check headers
        expect(page.locator('[aria-label="Длительность коммуникаций"]')).to_have_count(2)
        expect(page.locator('[aria-label="Время первой коммуникации"]')).to_have_count(2)
        expect(page.locator('[aria-label="Время последней коммуникации"]')).to_have_count(2)
        # check sum
        expect(page.locator('[title="80:48:49"]')).to_have_count(2)
        expect(page.locator('[title="08.02.2022, 00:12:20"]')).to_have_count(4)
        expect(page.locator('[title="16.05.2022, 18:21:38"]')).to_have_count(4)



@pytest.mark.e2e
@pytest.mark.reports
@allure.title("test_reports_diff")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_reports_diff")
def test_reports_diff(base_url, page: Page) -> None:
    reports = Reports(page)

    with allure.step("Go to url"):
        reports.navigate(base_url)

    with allure.step("Auth with ecotelecom"):
        reports.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to reports"):
        reports.click_reports()

    with allure.step("Press (Create report)"):
        reports.press_create_report()

    with allure.step("Choose period 10/02/2022-10/02/2022"):
        reports.choose_period_date("10/02/2022", "10/02/2022")

    with allure.step("Generate report"):
        reports.press_generate_report()

    with allure.step("Click diff report"):
        page.get_by_text("Настройка таблицы").click()
        page.wait_for_selector('[class="shown"]')
        page.locator('[data-testid="show_diff_report"]').locator('[type="checkbox"]').click()

    with allure.step("Check"):
        expect(page.locator('[title="419 (-1673)"]')).to_have_count(3)


