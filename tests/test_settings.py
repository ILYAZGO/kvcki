from playwright.sync_api import Page, expect, Route
from openpyxl import load_workbook
from utils.variables import *
from pages.settings import *
from utils.dates import *
from utils.create_delete_user import create_user, delete_user, give_users_to_manager, create_operator, give_access_right
import os
import pytest
import allure
import random


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_address_book_fill_by_user")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("input text to address-book, save, check that text saved")
def test_address_book_fill_by_user(base_url, page: Page) -> None:
    settings = Settings(page)

    text = "Телефон;Должность;ФИО\n12345;IVAN;BOSS"
    
    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)
    
    with allure.step("Go to url"):
        settings.navigate(base_url)
    
    with allure.step("Auth as user"):
        settings.auth(LOGIN, PASSWORD)
    
    with allure.step("Go to settings"):
        settings.click_settings()
    
    with allure.step("Go to address book"):
        settings.click_address_book()
    
    with allure.step("Fill address book"):
        settings.fill_address_book(text)
    
    with allure.step("Press (Save) button"):
        settings.press_save()
    
    with allure.step("Go to personal info"):
        settings.click_personal_info()
    
    with allure.step("Go to address book"):
        settings.click_address_book()
    
    with allure.step("Check that address book contain correct text"):
        settings.assert_address_book_text(text)
    
    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that address book contain correct text after page reload"):
        settings.assert_address_book_text(text)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_can_change_login_for_manager")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("check admin can change login for manager")
def test_admin_can_change_login_for_manager(base_url, page: Page) -> None:
    settings = Settings(page)

    CHANGED_LOGIN = f"auto_test_user_ch_{datetime.now().strftime('%m%d%H%M')}{datetime.now().microsecond}"
    
    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)
    
    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Go to url"):
        settings.navigate(base_url)
    
    with allure.step("Auth as admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)
    
    with allure.step("Go to manager"):
        settings.go_to_user(LOGIN_MANAGER)
    
    with allure.step("Go to settings"):
        settings.click_settings()
    
    with allure.step("Check that manager have original login"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(LOGIN_MANAGER)
    
    with allure.step("Change login"):
        settings.change_login(CHANGED_LOGIN)
    
    with allure.step("Press (save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Reload page"):
        settings.reload_page()
    
    with allure.step("Check that login changed"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(CHANGED_LOGIN)
    
    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)
    
    with allure.step("Delete manager"):    
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_can_change_login_for_user_and_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("check admin can change login for user and operator")
def test_admin_can_change_login_for_user_and_operator(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_OPERATOR_LOGIN = f"auto_test_operator_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    CHANGED_LOGIN = f"auto_test_user_ch_{datetime.now().strftime('%m%d%H%M')}{datetime.now().microsecond}"

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)
    
    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)
    
    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)
    
    with allure.step("Go to url"):
        settings.navigate(base_url)
    
    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)
    
    # change login for user
    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)
    
    with allure.step("Go to settings"):
        settings.click_settings()
    
    with allure.step("Change login for user"):
        settings.change_login(CHANGED_LOGIN)
    
    with allure.step("Press (save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Reload page"):
        settings.reload_page()
    
    with allure.step("Check that login changed"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(CHANGED_LOGIN)

    # change for operator
    with allure.step("Go to employees from left menu"):
        settings.click_employees()
    
    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()
    
    with allure.step("Change login for operator"):
        settings.change_login(NEW_OPERATOR_LOGIN)
    
    with allure.step("Press (save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Page reload"):
        settings.reload_page()
    
    with allure.step("Check that login changed"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(NEW_OPERATOR_LOGIN)
    
    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)
    
    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)
    
    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_cant_change_login_for_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("check that changing login disabled for operator")
def test_user_cant_change_login_for_operator(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Crete operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    # check disabled for operator
    with allure.step("Go to employees"):
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Check that login field in operators personal info disabled"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_manager_cant_change_login_for_user_and_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_manager_cant_change_login_for_user_and_operator")
def test_manager_cant_change_login_for_user_and_operator(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)
        give_users_to_manager(API_URL, USER_ID_MANAGER, [USER_ID_USER, importFrom_user_id], TOKEN_MANAGER)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with manager"):
        settings.auth(LOGIN_MANAGER, PASSWORD)

    with allure.step("Go to user"):
        # check disabled for user
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that login field disabled"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()

    with allure.step("Go to employees"):
        # check disabled for operator
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Check that login field disabled"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()

    with allure.step("Delete manager"):
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_can_change_rights_for_manager")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("check admin can change rights for manager")
def test_admin_can_change_rights_for_manager(base_url, page: Page) -> None:
    settings = Settings(page)
    
    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)
    
    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)
    
    with allure.step("Go to url"):
        settings.navigate(base_url)

    with allure.step("Auth as admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)
    
    with allure.step("Go to manager"):
        settings.go_to_user(LOGIN_MANAGER)
    
    with allure.step("Go to settings"):
        settings.click_settings()
    
    with allure.step("Go to rights"):
        settings.click_rights()
    
    with allure.step("Check that manager have 6 right"):
        expect(page.locator(BLOCK_ONE_RIGHT)).to_have_count(7)
    
    with allure.step("Click to all rights and check all checkboxes"):
        settings.click_all_checkboxes_on_page()
    
    with allure.step("Press (save) in rights"):
        settings.press_save_in_rights()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Page reload"):
        settings.reload_page()
    
    with allure.step("Check that all rights checked after page reload"):
        all_checkboxes_to_be_checked(page)

        assert all_checkboxes_to_be_checked(page) == True
    
    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)
    
    with allure.step("Delete manager"):
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_can_change_rights_for_user_and_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("check admin can change rights for user and operator")
def test_admin_can_change_rights_for_user_and_operator(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)
    
    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)
    
    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to url"):
        settings.navigate(base_url)

    with allure.step("Auth as admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    # change rights for user
    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)
    
    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to rights"):
        settings.click_rights()
    
    with allure.step("Check that user have 3 right"):
        expect(page.locator(BLOCK_ONE_RIGHT)).to_have_count(3)
    
    with allure.step("Click to all rights and check all checkboxes"):
        settings.click_all_checkboxes_on_page()
    
    with allure.step("Press (save) in rights"):
        settings.press_save_in_rights()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Page reload"):
        settings.reload_page()
    
    with allure.step("Check that all rights checked after page reload"):
        all_checkboxes_to_be_checked(page)

        assert all_checkboxes_to_be_checked(page) == True

    # change rights for operator
    
    with allure.step("Go to employees"):
        settings.click_employees()
    
    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()
    
    with allure.step("Go to rights"):
        settings.click_rights()
    
    with allure.step("Check that operator have 27 rights"):
        expect(page.locator(BLOCK_ONE_RIGHT)).to_have_count(29)
    
    with allure.step("Click to all rights and check all checkboxes"):
        settings.click_all_checkboxes_on_page()
    
    with allure.step("Press (save) in rights"):
        settings.press_save_in_rights()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")
    
    with allure.step("Page reload"):
        settings.reload_page()
    
    with allure.step("Check that all rights checked after page reload"):
        all_checkboxes_to_be_checked(page)

        assert all_checkboxes_to_be_checked(page) == True
    
    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)
    
    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)
    
    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_can_change_rights_for_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check user can change rights for operator")
def test_user_can_change_rights_for_operator(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    # change rights for operator
    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to employees"):
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Go to rights"):
        settings.click_rights()

    with allure.step("Check that operator have 27 rights in list"):
        expect(page.locator(BLOCK_ONE_RIGHT)).to_have_count(29)

    with allure.step("Click to all rights and check all checkboxes"):
        settings.click_all_checkboxes_on_page()

    with allure.step("Press (Save) button"):
        settings.press_save_in_rights()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that all rights checked after page reload"):
        all_checkboxes_to_be_checked(page)

        assert all_checkboxes_to_be_checked(page) == True

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_admin_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for admin")
def test_change_personal_information_save_admin_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_NAME = NEW_LOGIN = f"auto_test_user_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    EMAIL = f"email_{datetime.now().microsecond}{random.randint(100, 200)}@mail.ru"

    with allure.step("Create admin"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Change login"):
        settings.change_login(NEW_LOGIN)

    with allure.step("Press (Save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Change person al information"):
        settings.fill_personal_information_admin_and_manager(
            NEW_NAME,
            EMAIL,
            "1234567890",
            "someComment",
            "Africa/Bamako",
            "PT"
        )

    with allure.step("Press (Save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Click notifications"):
        settings.click_notifications()

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that personal information changed"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(NEW_LOGIN)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that personal information still have after reboot"):
        expect(page.locator(INPUT_LOGIN)).to_have_value(NEW_LOGIN)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_manager_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for manager")
def test_change_personal_information_save_manager_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_NAME = f"auto_test_user_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    EMAIL = f"email_{datetime.now().microsecond}{random.randint(100, 999)}@mail.ru"

    with allure.step("Create manager"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Change personal information"):
        settings.fill_personal_information_admin_and_manager(
            NEW_NAME,
            EMAIL,
            "1234567890",
            "someComment",
            "Africa/Bamako",
            "PT"
        )

    with allure.step("Press (Save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Click notifications"):
        settings.click_notifications()

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that personal information changed"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that personal information still have after reboot"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_user_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for user")
def test_change_personal_information_save_user_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_NAME = f"auto_test_user_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    EMAIL = f"email_{datetime.now().microsecond}{random.randint(100, 999)}@mail.ru"

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Change person al information"):
        settings.fill_personal_information_user_and_operator(
            NEW_NAME,
            EMAIL,
            "1234567890",
            "Africa/Bamako",
            "PT"
        )

    with allure.step("Press (Save) button"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Click notifications"):
        settings.click_notifications()

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that personal information changed"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).not_to_be_visible()
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()
        expect(page.locator(INPUT_NEW_PASSWORD)).to_be_enabled()
        expect(page.locator(INPUT_NEW_PASSWORD_REPEAT)).to_be_disabled()

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that personal information still have after reboot"):
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).not_to_be_visible()
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()
        expect(page.locator(INPUT_NEW_PASSWORD)).to_be_enabled()
        expect(page.locator(INPUT_NEW_PASSWORD_REPEAT)).to_be_disabled()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_operator_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for operator")
def test_change_personal_information_save_operator_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    message = "Редактировать персональную информацию может только администратор"

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with operator"):
        settings.auth(LOGIN_OPERATOR, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that operator cant change personal information"):
        expect(page.locator(INPUT_EMAIL)).to_be_disabled()
        expect(page.locator(INPUT_PHONE)).to_be_disabled()
        expect(page.locator(INPUT_COMMENT)).not_to_be_visible()
        expect(page.locator(INPUT_LOGIN)).to_be_disabled()
        expect(page.locator(INPUT_NAME)).to_be_disabled()
        expect(page.locator(INPUT_NEW_PASSWORD)).to_be_disabled()
        expect(page.locator(INPUT_NEW_PASSWORD_REPEAT)).to_be_disabled()
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Check that message (Can change only administrator) exists"):
        expect(page.locator('[class*="typography--variant-body2"]')).to_contain_text(message)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_operator_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for operator by admin")
def test_change_personal_information_save_operator_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_OPERATOR_NAME = NEW_OPERATOR_LOGIN = f"auto_test_operator_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    EMAIL = f"email_{datetime.now().microsecond}{random.randint(100, 999)}@mail.ru"

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to employees"):
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Fill personal information"):
        settings.fill_personal_information_admin_and_manager(
            NEW_OPERATOR_NAME,
            EMAIL,
            "1234567890",
            "someComment",
            "Africa/Bamako",
            "PT"
        )

    with allure.step("Press (save)"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Go to rights"):
        settings.click_rights()

    with allure.step("Go to personal info"):
        settings.click_personal_info()

    with allure.step("Check that personal information saved"):
        expect(page.locator(BLOCK_PERSONAL_INFO)).not_to_contain_text("Редактировать ")
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_OPERATOR_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that personal information saved"):
        expect(page.locator(BLOCK_PERSONAL_INFO)).not_to_contain_text("Редактировать ")
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_OPERATOR_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).to_have_value("someComment")
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_personal_information_save_operator_by_user")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check changing and saving personal info for operator by user")
def test_change_personal_information_save_operator_by_user(base_url, page: Page) -> None:
    settings = Settings(page)

    NEW_OPERATOR_NAME = NEW_OPERATOR_LOGIN = f"auto_test_operator_{datetime.now().strftime('%m%d%H%M')}_{datetime.now().microsecond}"
    EMAIL = f"email_{datetime.now().microsecond}{random.randint(100, 999)}@mail.ru"

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to employees"):
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Fill personal information"):
        settings.fill_personal_information_user_and_operator(
            NEW_OPERATOR_NAME,
            EMAIL,
            "1234567890",
            "Africa/Bamako",
            "PT"
        )

    with allure.step("Press (save)"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Go to rights"):
        settings.click_rights()

    with allure.step("Go to personal info"):
        settings.click_personal_info()

    with allure.step("Check that personal information saved"):
        expect(page.locator(BLOCK_PERSONAL_INFO)).not_to_contain_text("Редактировать ")
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_OPERATOR_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).not_to_be_visible()
        expect(page.locator(SELECT_TIMEZONE)).to_have_text("Africa/Bamako", timeout=wait_until_visible)
        expect(page.locator(SELECT_USER_LANG)).to_have_text("PT", timeout=wait_until_visible)
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that personal information saved"):
        expect(page.locator(BLOCK_PERSONAL_INFO)).not_to_contain_text("Редактировать ")
        expect(page.locator(INPUT_NAME)).to_have_value(NEW_OPERATOR_NAME)
        expect(page.locator(INPUT_EMAIL)).to_have_value(EMAIL)
        expect(page.locator(INPUT_PHONE)).to_have_value("1234567890")
        expect(page.locator(INPUT_COMMENT)).not_to_be_visible()
        expect(page.locator(SELECT_TIMEZONE).get_by_text("Africa/Bamako")).to_be_visible()
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_left_menu_items_for_admin_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check how many items in left menu for role")
def test_left_menu_items_for_admin_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check items in left menu"):
        expect(page.locator(BLOCK_LEFT_MENU)).to_contain_text(['Персональная информация'])
        expect(page.locator(LEFT_MENU_ITEM)).to_have_count(1)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN, USER_ID)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_left_menu_items_for_manager_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check how many items in left menu for role")
def test_left_menu_items_for_manager_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create manager"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with manager"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check items in left menu"):
        expect(page.locator(BLOCK_LEFT_MENU)).to_contain_text(['Персональная информация'])
        expect(page.locator(LEFT_MENU_ITEM)).to_have_count(1)

    with allure.step("Delete manager"):
        delete_user(API_URL, TOKEN, USER_ID)


# check how many items in left menu for role
@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_left_menu_items_for_user_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check how many items in left menu for role")
def test_left_menu_items_for_user_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    block_list = "Персональная информацияСотрудникиДействия с коммуникациямиКвоты776История потребления услугТарифыАдресная книгаИнтеграции"

    with allure.step("Create user"):
        USER_ID, TOKEN, LOGIN = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check items in left menu"):
        page.wait_for_timeout(1500)
        expect(page.locator(BLOCK_LEFT_MENU)).to_contain_text(block_list)
        expect(page.locator(LEFT_MENU_ITEM)).to_have_count(8)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN, USER_ID)

@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_left_menu_items_for_operator_itself")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("Check how many items in left menu for role")
def test_left_menu_items_for_operator_itself(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL,USER_ID_USER,PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with operator"):
        settings.auth(LOGIN_OPERATOR, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check items in left menu"):
        expect(page.locator(BLOCK_LEFT_MENU)).to_contain_text(['Персональная информация'])
        expect(page.locator(LEFT_MENU_ITEM)).to_have_count(1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_check_industry_and_partner_for_manager")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_admin_check_industry_and_partner_for_manager")
def test_admin_check_industry_and_partner_for_manager(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to manager"):
        settings.go_to_user(LOGIN_MANAGER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that industry and partner not visible"):
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete mamager"):
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_admin_check_industry_and_partner_for_user_and_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_admin_check_industry_and_partner_for_user_and_operator")
def test_admin_check_industry_and_partner_for_user_and_operator(base_url, page: Page) -> None:
    industys_list = ("Не выбраноАвтомобильная индустрияB2B продажиE-commerceМедклиникаСтоматологияКолл-центрыКонсалтинг"
                     "ЛогистикаМебельНефть и газНедвижимостьПроизводствоРозничные продажиРазработка ПОСфера услугТуризм"
                     "ФинансыФитнес-центрыEdTechТелекоммуникацииHR и рекрутинг")

    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        # check and change for user
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that industry and partner visible"):
        expect(page.locator(SELECT_INDUSTRY)).to_be_visible()
        expect(page.locator(SELECT_PARTNER)).to_be_visible()

    with allure.step("Check industrys list"):
        settings.assert_industries_list(industys_list)

    with allure.step("Change industry"):
        settings.change_industry('E-commerce')

    with allure.step("Change partner"):
        settings.change_partner('managerIM')

    with allure.step("Press (save)"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that industry and partner changed"):
        expect(page.locator(SELECT_INDUSTRY)).to_have_text('E-commerce')
        expect(page.locator(SELECT_PARTNER)).to_have_text('managerIM')

    with allure.step("Go to employees"):
        # check for operator
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Check that industry and partner NOT visible"):
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_manager_check_industry_and_partner_for_user_and_operator")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_manager_check_industry_and_partner_for_user_and_operator")
def test_manager_check_industry_and_partner_for_user_and_operator(base_url, page: Page) -> None:
    industys_list = ("Не выбраноАвтомобильная индустрияB2B продажиE-commerceМедклиникаСтоматологияКолл-центрыКонсалтинг"
                     "ЛогистикаМебельНефть и газНедвижимостьПроизводствоРозничные продажиРазработка ПОСфера услугТуризм"
                     "ФинансыФитнес-центрыEdTechТелекоммуникацииHR и рекрутинг")

    settings = Settings(page)

    with allure.step("Create manager"):
        USER_ID_MANAGER, TOKEN_MANAGER, LOGIN_MANAGER = create_user(API_URL, ROLE_MANAGER, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)
        give_users_to_manager(API_URL, USER_ID_MANAGER, [USER_ID_USER, importFrom_user_id], TOKEN_MANAGER)

    with allure.step("Create operator"):
        USER_ID_OPERATOR, TOKEN_OPERATOR, LOGIN_OPERATOR = create_operator(API_URL, USER_ID_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with manager"):
        settings.auth(LOGIN_MANAGER, PASSWORD)

    with allure.step("Go to user"):
        # check and change for user
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Check that industry visible and partner NOT visible"):
        expect(page.locator(SELECT_INDUSTRY)).to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Check industrys list"):
        settings.assert_industries_list(industys_list)

    with allure.step("Change industry"):
        settings.change_industry('E-commerce')

    with allure.step("Press (save)"):
        settings.press_save()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Профиль успешно сохранен")

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check that industry changed and partner NOT visible"):
        expect(page.locator(SELECT_INDUSTRY)).to_have_text('E-commerce')
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Go to employees"):
        # check for operator
        settings.click_employees()

    with allure.step("Go to operator from table"):
        settings.go_to_operator_from_table()

    with allure.step("Check that industry and partner NOT visible"):
        expect(page.locator(SELECT_INDUSTRY)).not_to_be_visible()
        expect(page.locator(SELECT_PARTNER)).not_to_be_visible()

    with allure.step("Delete manager"):
        delete_user(API_URL, TOKEN_MANAGER, USER_ID_MANAGER)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

    with allure.step("Delete operator"):
        delete_user(API_URL, TOKEN_OPERATOR, USER_ID_OPERATOR)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_giving_communications_quota_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("every auto_test_user gets 777 min quota by default. test check that we can add more")
def test_giving_communications_quota_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to quotas"):
        settings.click_quota()

    with allure.step("Click (add quota)"):
        settings.press_add_in_quotas()

    with allure.step("Check that system will recommend last gived quota (777)"):
        expect(page.locator(INPUT_QUOTA_TIME)).to_have_value("777")

    with allure.step("Click checkbox (bessro4no)"):
        page.locator(MODAL_WINDOW).locator('[type="checkbox"]').click()

    with allure.step("Check that checkbox was checked"):
        expect(page.locator(MODAL_WINDOW).locator('[type="checkbox"]')).to_be_checked()

    with allure.step("Check that dates disabled"):
        expect(page.locator('[class*="ant-picker-disabled"]')).to_be_visible()

    with allure.step("Fill quota 100"):
        settings.fill_quota_time("100")

    with allure.step("press (add)"):
        settings.press_add_in_modal_in_quotas()
        #page.wait_for_selector('[aria-rowindex="2"]', timeout=wait_until_visible)

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Квота добавлена")

    with allure.step("Reload page"):
        page.reload()
        page.wait_for_selector('[aria-rowindex="2"]', timeout=wait_until_visible)
        page.wait_for_timeout(2000)

    with allure.step("Check that quota added"):
        expect(page.locator('[role="gridcell"]')).to_have_count(12)
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_have_text("Бессрочно")

    with allure.step("Delete added quota"):
        page.locator('[role="grid"]').locator('[fill="#FF4D4F"]').click()
        page.wait_for_timeout(500)

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Квота удалена")

    with allure.step("Check that quota deleted"):
        expect(page.locator('[role="grid"]').locator('[fill="#FF4D4F"]')).not_to_be_visible(timeout=wait_until_visible)
        #expect(page.locator('[class="rs-table-body-info"]')).to_have_text("Информация отсутствует")

    with allure.step("Click (add quota)"):
        settings.press_add_in_quotas()

    with allure.step("Check that system will recommend last gived quota (100)"):
        expect(page.locator(INPUT_QUOTA_TIME)).to_have_value("100")

    with allure.step("Choose period for quota"):
        settings.choose_period_date(tomorrow, tomorrow)

    with allure.step("Fill quota 100"):
        settings.fill_quota_time("100")

    with allure.step("press (add)"):
        settings.press_add_in_modal_in_quotas()
        #page.wait_for_selector('[aria-rowindex="2"]', timeout=wait_until_visible)

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Квота добавлена")

    with allure.step("Reload page"):
        page.reload()
        page.wait_for_selector('[aria-rowindex="2"]', timeout=wait_until_visible)
        page.wait_for_timeout(2000)

    with allure.step("Check that quota added"):
        expect(page.locator('[role="gridcell"]')).to_have_count(18)
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_have_text(f"{tomorrow_with_dash} - {tomorrow_with_dash}")

    with allure.step("Delete added quota"):
        page.locator('[fill="#FF4D4F"]').click()
        page.wait_for_timeout(500)

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Квота удалена")

    with allure.step("Check that quota deleted"):
        expect(page.locator('[fill="#FF4D4F"]')).not_to_be_visible(timeout=wait_until_visible)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_giving_gpt_quota_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("every auto_test_user gets 777 min quota by default. test check that we can add more")
def test_giving_gpt_quota_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to quotas"):
        settings.click_quota()

    with allure.step("Click to (GPT) tab"):
        settings.click_to_gpt_tab()

        # gpt
    with allure.step("Check negative value"):
        settings.fill_quota_value(0, "-1")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(0, "151")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(0, "150")

    with allure.step("Click (save)"):
        page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT).click()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Данные успешно обновлены")

    with allure.step("Reload page and check that saved and have residue"):
        settings.reload_page()
        expect(page.get_by_text("Ограничение 150 USD")).to_have_count(1, timeout=wait_until_visible)
        expect(page.get_by_text("150.00")).to_have_count(2)

        # yandex
    with allure.step("Check negative value"):
        settings.fill_quota_value(1, "-1")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(1, "15001")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(1, "15000")

    with allure.step("Click (save)"):
        page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT).click()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Данные успешно обновлены")

    with allure.step("Reload page and check that saved and have residue"):
        settings.reload_page()
        expect(page.get_by_text("Ограничение 15000 RUB")).to_have_count(1, timeout=wait_until_visible)
        expect(page.get_by_text("15000.00")).to_have_count(2)

        # imotio
    with allure.step("Check negative value"):
        settings.fill_quota_value(2, "-1")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(2, "1000001")

    with allure.step("Check that (save) button disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        page.wait_for_timeout(500)

    with allure.step("Check value more than limit"):
        settings.fill_quota_value(2, "1000000")

    with allure.step("Click (save)"):
        page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT).click()

    with allure.step("Wait for snackbar and check"):
        settings.check_alert("Данные успешно обновлены")

    with allure.step("Reload page and check that saved and have residue"):
        settings.reload_page()
        expect(page.get_by_text("Ограничение 1000000 RUB")).to_have_count(1, timeout=wait_until_visible)
        expect(page.get_by_text("1000000.00")).to_have_count(2)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_giving_gpt_quota_by_admin_if_500")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_giving_gpt_quota_by_admin_if_500. if /gpt_quotas 500")
def test_giving_gpt_quota_by_admin_if_500(base_url, page: Page) -> None:
    settings = Settings(page)

    error = "Данные о GPT квотах не получены"

    def handle_consumption(route: Route):
        route.fulfill(status=500, body="")
    # Intercept the route
    page.route("**/gpt_quotas", handle_consumption)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to quotas"):
        settings.click_quota()

    with allure.step("Click to (GPT) tab"):
        settings.click_to_gpt_tab()

    with allure.step("Check error"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_cant_change_quotas")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("every auto_test_user gets 777 min quota by default. test check that we can add more")
def test_user_cant_change_quotas(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to quotas"):
        settings.click_quota()

    with allure.step("Check that button (Add) not visible for user"):
        expect(page.get_by_role("button", name="Добавить", exact=True)).not_to_be_visible()

    with allure.step("Click to (GPT) tab"):
        settings.click_to_gpt_tab()

    with allure.step("Check that button (save) and input for new amount - disabled"):
        expect(page.locator(BLOCK_WITH_SAVE_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()
        expect(page.locator('[placeholder="Новое значение"]').nth(0)).to_be_disabled()
        expect(page.locator('[placeholder="Новое значение"]').nth(1)).to_be_disabled()
        expect(page.locator('[placeholder="Новое значение"]').nth(2)).to_be_disabled()

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_consumption_history")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_consumption_history. mocked")
def test_user_consumption_history(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    # mocks
    def handle_calls(route: Route):
        json_calls = [
            {"callDate": "2024-07-28", "fromServices": ["service1"], "callCount": 111, "sumDuration": 395},
            {"callDate": "2024-07-29", "fromServices": ["service2"], "callCount": 222, "sumDuration": 678},
            {"callDate": "2024-07-30", "fromServices": [], "callCount": 444, "sumDuration": 777}
        ]
        # fulfill the route with the mock data
        route.fulfill(json=json_calls)
        # Intercept the route
    page.route("**/history/calls?**", handle_calls)

    def handle_chats(route: Route):
        json_calls = [
            {"chatDate": "2024-07-28", "fromServices": [""], "chatCount": 100},
            {"chatDate": "2024-07-29", "fromServices": ["chat1"], "chatCount": 999},
            {"chatDate": "2024-07-30", "fromServices": ["chat2"], "chatCount": 9002}

        ]
        # fulfill the route with the mock data
        route.fulfill(json=json_calls)
        # Intercept the route
    page.route("**/history/chats?**", handle_chats)

    def handle_gpt(route: Route):
        json = [
            {"gptDate": "2024-07-30", "engine": "yandex_gpt", "model": "yandexgpt-lite", "communicationType": "call",
             "gptRequestId": "66A8B7129D963878D93D73CD", "gptRequestTitle": "title1", "requestCount": "200",
             "communicationCount": "111", "totalAmmount": "7.468260013125836"},
            {"gptDate": "2024-07-29", "engine": "chat_gpt", "model": "gpt-3.5-turbo", "communicationType": "call",
             "gptRequestId": "66A8B7129D963878D93D73CF", "gptRequestTitle": "title2", "requestCount": "299",
             "communicationCount": "222", "totalAmmount": "7.418260013125836"}
        ]
        # fulfill the route with the mock data
        route.fulfill(json=json)
        # Intercept the route
    page.route("**/history/gpt?**", handle_gpt)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_CONSUMPTION_HISTORY).click()

    with allure.step("Check existing search for audio, calendar, mocked data and total count"):
        expect(page.locator(SEARCH_IN_CONSUMPTION_AUDIO)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)
        #  check first row
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="1"]')).to_contain_text("28.07.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="2"]')).to_contain_text("service1")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="3"]')).to_contain_text("111")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_contain_text("395")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="5"]')).to_contain_text("6:35")
        #  check second row
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="1"]')).to_contain_text("29.07.2024")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="2"]')).to_contain_text("service2")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="3"]')).to_contain_text("222")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="4"]')).to_contain_text("678")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="5"]')).to_contain_text("11:18")
        #  check third row
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="1"]')).to_contain_text("30.07.2024")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="2"]')).to_contain_text("")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="3"]')).to_contain_text("444")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="4"]')).to_contain_text("777")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="5"]')).to_contain_text("12:57")
        #  check total count
        expect(page.locator(TOTAL_AUDIO_MIN)).to_contain_text("1850")
        expect(page.locator(TOTAL_AUDIO_HOURS)).to_contain_text("30:50")

    with allure.step("Fill search by service1"):
        page.locator(SEARCH_IN_CONSUMPTION_AUDIO).locator('[type="text"]').fill("service1")

    with allure.step("Check search that service2 not visible"):
        expect(page.locator('[class*="communicationsStyles_table_"]')).not_to_contain_text("service2")

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_GPT).click()

    with allure.step("Check exist search for chats, calendar, mocked data and total count"):
        expect(page.locator(SEARCH_IN_CONSUMPTION_GPT)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по движку, модели, типу коммуникации или запросу"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)
        #  check first row
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="1"]')).to_contain_text("30.07.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="2"]')).to_contain_text("yandex_gpt")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="3"]')).to_contain_text("yandexgpt-lite")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_contain_text("call")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="5"]')).to_contain_text("title1")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="6"]')).to_contain_text("200")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="7"]')).to_contain_text("111")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="8"]')).to_contain_text("7.47")
        #  check second row
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="1"]')).to_contain_text("29.07.2024")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="2"]')).to_contain_text("chat_gpt")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="3"]')).to_contain_text("gpt-3.5-turbo")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="4"]')).to_contain_text("call")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="5"]')).to_contain_text("title2")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="6"]')).to_contain_text("299")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="7"]')).to_contain_text("222")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="8"]')).to_contain_text("7.42")
        #  check total count
        expect(page.locator(TOTAL_GPT_MONEY)).to_contain_text("14.89")

    with allure.step("Fill search by ya"):
        page.locator(SEARCH_IN_CONSUMPTION_GPT).locator('[type="text"]').fill("ya")

    with allure.step("Check search that chat_gpt not visible"):
        expect(page.locator('[class*="communicationsStyles_table_"]')).not_to_contain_text("chat_gpt")

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_CHATS).click()

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(SEARCH_IN_CONSUMPTION_CHATS)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)
        #  check first row
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="1"]')).to_contain_text("28.07.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="2"]')).to_contain_text("")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="3"]')).to_contain_text("100")
        #  check second row
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="1"]')).to_contain_text("29.07.2024")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="2"]')).to_contain_text("chat1")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="3"]')).to_contain_text("999")
        #  check third row
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="1"]')).to_contain_text("30.07.2024")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="2"]')).to_contain_text("chat2")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="3"]')).to_contain_text("9002")
        #  check total count
        expect(page.locator(TOTAL_CHATS)).to_contain_text("10101")

    with allure.step("Fill search by chat1"):
        page.locator(SEARCH_IN_CONSUMPTION_CHATS).locator('[type="text"]').fill("chat1")

    with allure.step("Check search that chat_gpt not visible"):
        expect(page.locator('[class*="communicationsStyles_table_"]')).not_to_contain_text("chat2")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_consumption_history_if_empty")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_consumption_history. check have warning if []")
def test_user_consumption_history_if_empty(base_url, page: Page) -> None:
    settings = Settings(page)

    alert_massage = "Нет информации об истории потребления за выбранный период"

    def handle_consumption(route: Route):
        route.fulfill(body="[]")
    # Intercept the route
    page.route("**/history/calls?**", handle_consumption)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_CONSUMPTION_HISTORY).click()
        page.wait_for_timeout(2000)
        page.wait_for_selector(CONSUMPTION_ERROR_FIRST_LINE)

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(alert_massage)
        expect(page.locator(SEARCH_IN_CONSUMPTION_AUDIO)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_disabled()
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_GPT).click()

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(alert_massage)
        expect(page.locator(SEARCH_IN_CONSUMPTION_GPT)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по движку, модели, типу коммуникации или запросу"]')).to_have_count(1)
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_disabled()
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_CHATS).click()

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(alert_massage)
        expect(page.locator(SEARCH_IN_CONSUMPTION_CHATS)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_disabled()
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_consumption_history_if_500")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_consumption_history_if_500. check have warning if 500")
def test_user_consumption_history_if_500(base_url, page: Page) -> None:
    settings = Settings(page)

    error_first_line = "Возникла ошибка в процессе формирования таблицы"
    error_second_line = "Пожалуйста, перезагрузите страницу браузера или зайдите в систему заново"

    def handle_consumption(route: Route):
        route.fulfill(status=500, body="")
    # Intercept the route
    page.route("**/history/calls?**", handle_consumption)
    page.route("**/history/chats?**", handle_consumption)
    page.route("**/history/gpt?**", handle_consumption)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_CONSUMPTION_HISTORY).click()
        page.wait_for_timeout(2000)
        page.wait_for_selector(CONSUMPTION_ERROR_FIRST_LINE)

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error_first_line)
        expect(page.locator(CONSUMPTION_ERROR_SECOND_LINE)).to_contain_text(error_second_line)
        expect(page.locator(SEARCH_IN_CONSUMPTION_AUDIO)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_GPT).click()

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error_first_line)
        expect(page.locator(CONSUMPTION_ERROR_SECOND_LINE)).to_contain_text(error_second_line)
        expect(page.locator(SEARCH_IN_CONSUMPTION_GPT)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по движку, модели, типу коммуникации или запросу"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_CHATS).click()

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error_first_line)
        expect(page.locator(CONSUMPTION_ERROR_SECOND_LINE)).to_contain_text(error_second_line)
        expect(page.locator(SEARCH_IN_CONSUMPTION_CHATS)).to_have_count(1)
        expect(page.locator('[placeholder="Поиск по источнику"]')).to_have_count(1)
        expect(page.locator(CALENDAR_IN_CONSUMPTION)).to_have_count(1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

#####################
@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_consumption_history_export")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_consumption_history_export")
def test_user_consumption_history_export(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(ECOTELECOM, ECOPASS)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_CONSUMPTION_HISTORY).click()

    with allure.step("Check (EX) button is disabled"):
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_disabled()

    with allure.step("Input date"):
        settings.choose_period_date("03/04/2025", "03/04/2025")

    with allure.step("Click (EX)"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]').click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 5000 < os.path.getsize(path + download.suggested_filename) < 6000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Дата"
        assert sheet["A2"].value == "2025-04-03"
        assert sheet["A3"].value == "Итого за период"
        assert sheet["B1"].value == "Источник"
        assert sheet["B2"].value == "manual"
        assert sheet["C1"].value == "Количество коммуникаций"
        assert sheet["C2"].value == 1
        assert sheet["D1"].value == "Суммарная длительность в минутах"
        assert sheet["D2"].value == 2
        assert sheet["D3"].value == 2
        assert sheet["E1"].value == "Суммарная длительность в часах"
        assert sheet["E2"].value == "00:02"
        assert sheet["E3"].value == "00:02"
        assert sheet.max_row == 3
        assert sheet.max_column == 5

    with allure.step("Remove downloaded export"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_GPT).click()

    with allure.step("Check (EX) button is enabled"):
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_enabled()

    with allure.step("Input date"):
        settings.choose_period_date("03/04/2025", "03/04/2025")

    with allure.step("Click (EX)"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]').click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 6000 < os.path.getsize(path + download.suggested_filename) < 7000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Дата"
        assert sheet["A2"].value == "2025-04-03"
        assert sheet["A3"].value == "2025-04-03"
        assert sheet["A4"].value == "2025-04-03"
        assert sheet["A5"].value == "2025-04-03"
        assert sheet["A6"].value == "Итого за период"
        assert sheet["B1"].value == "Движок"
        assert sheet["B2"].value == "imotio_gpt"
        assert sheet["B3"].value == "chat_gpt"
        assert sheet["B4"].value == "imotio_gpt"
        assert sheet["B5"].value == "chat_gpt"
        assert sheet["C1"].value == "Модель"
        assert sheet["C2"].value == "auto"
        assert sheet["C3"].value == "auto"
        assert sheet["C4"].value == "auto"
        assert sheet["C5"].value == "gpt-3.5-turbo"
        assert sheet["D1"].value == "Тип коммуникации"
        assert sheet["D2"].value == "call"
        assert sheet["D3"].value == "call"
        assert sheet["D4"].value == "call"
        assert sheet["D5"].value == "call"
        assert sheet["E1"].value == "Запрос GPT"
        assert sheet["E2"].value == "Информация отсутствует"
        assert sheet["E3"].value == "Информация отсутствует"
        assert sheet["E4"].value == "Проверочное правило"
        assert sheet["E5"].value == "Информация отсутствует"
        assert sheet["F1"].value == "Количество запросов"
        assert sheet["F2"].value == 50
        assert sheet["F3"].value == 48
        assert sheet["F4"].value == 4
        assert sheet["F5"].value == 4
        assert sheet["G1"].value == "Количество коммуникаций"
        assert sheet["G2"].value == 2
        assert sheet["G3"].value == 1
        assert sheet["G4"].value == 1
        assert sheet["G5"].value == 1
        assert sheet["H1"].value == "Сумма (руб)"
        assert sheet["H2"].value == 0
        assert sheet["H3"].value == 0.06
        assert sheet["H4"].value == 0
        assert sheet["H5"].value == 0.12
        assert sheet["H6"].value == 0.18
        assert sheet.max_row == 6
        assert sheet.max_column == 8

    with allure.step("Remove downloaded export"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CONSUMPTION_HISTORY_CHATS).click()

    with allure.step("Check (EX) button is disabled"):
        expect(page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]')).to_be_disabled()

    with allure.step("Input date"):
        settings.choose_period_date("03/04/2025", "03/04/2025")

    with allure.step("Click (EX)"):
        # Start waiting for the download
        with page.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            page.locator('[aria-label="Excel экспорт"]').locator('[type="button"]').click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export downloaded"):
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 5000 < os.path.getsize(path + download.suggested_filename) < 6000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Дата"
        assert sheet["A2"].value == "2025-04-03"
        assert sheet["A3"].value == "Итого за период"
        assert sheet["B1"].value == "Источник"
        assert sheet["B2"].value == "usedesk"
        assert sheet["C1"].value == "Количество чатов"
        assert sheet["C2"].value == 1
        assert sheet["C3"].value == 1
        assert sheet.max_row == 3
        assert sheet.max_column == 3

    with allure.step("Remove downloaded export"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export removed"):
        assert os.path.isfile(path + download.suggested_filename) == False


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_check_word_processing_russian_language")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_check_word_processing_russian_language")
def test_check_word_processing_russian_language(base_url, page: Page) -> None:
    settings = Settings(page)

    # expected_languages = ("Английский (Великобритания)Английский (США)Испанский (Латинская Америка, Карибский регион, "
    #                       "код региона UN M49)Испанский (Испания)Французский (Франция)Португальский "
    #                       "(Бразилия)Португальский (Португалия)РусскийТурецкий (Турция)УкраинскийУзбекскийАвто")

    # expected_engines = "DeepgramGigaAMHappyscribeNLab SpeechIMOT.IOwhisperЯндексЯндекс v3"
    # expected_engines = "DeepgramgigaamHappyscribenexaraNLab SpeechIMOT.IOwhisperЯндексyandex_v3"

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to word processing"):
        settings.click_word_processing()

    with allure.step("Check default language"):
        expect(page.locator(SELECT_LANGUAGE)).to_contain_text("Русский")

    with allure.step("Click to language"):
        settings.click_language_select()

    with allure.step("Check language list"):
        expect(page.locator(MENU)).to_contain_text(fucking_stupidity)

    with allure.step("Close language menu"):
        page.locator('[class*="STT_order_"]').click()

    with allure.step("Check that engine not selected"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("NLab Speech")

    with allure.step("Check that model not selected"):
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Check engine list"):
        expect(page.locator(MENU)).to_contain_text(expected_engines)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Deepgram"):
        settings.choose_option(1)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Обобщённаяwhisper")

    with allure.step("Select model Обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model whisper"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Deepgram")
        expect(page.locator(SELECT_MODEL)).to_contain_text("whisper")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Happyscribe"):
        settings.choose_option(3)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Happyscribe")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose NLab Speech"):
        settings.choose_option(4)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("ОбобщённаяЖадный")

    with allure.step("Select model Обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Жадный"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("NLab Speech")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Жадный")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose IMOT.IO"):
        settings.choose_option(5)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("IMOT.IO")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose whisper"):
        settings.choose_option(6)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("whisper")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(5)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Яндекс"):
        settings.choose_option(7)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Обобщённая"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)
#
    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose yandex_v3"):
        settings.choose_option(8)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Обобщённая"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс v3")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ENGINE_DIARIZATION)).to_be_checked()
        expect(page.locator(CHECKBOX_NORMALIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PROFANITY_FILTER)).not_to_be_checked()
        expect(page.locator(CHECKBOX_LITERATURE_STYLE)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PHONE_FORMATTING)).to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(8)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_check_word_processing_parameters_combination")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_check_word_processing_parameters_combination")
def test_check_word_processing_parameters_combination(base_url, page: Page) -> None:
    settings = Settings(page)

    alert_merge = "Опция 'Объединить дорожки в один файл' не может быть выбрана одновременно с любой из диаризаций"

    alert_diarization = "Выберите только одну диаризацию среди опций"

    data_updated = "Данные успешно обновлены"

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to word processing"):
        settings.click_word_processing()

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose assembly_ai"):
        settings.choose_option(8)

    with allure.step("Check (Save) button is disabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_disabled()

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model best"):
        settings.choose_option(0)

    with allure.step("Check (Save) button is enabled"):
        expect(page.locator(BLOCK_WITH_BUTTON).locator(BUTTON_SUBMIT)).to_be_enabled()

    with allure.step("Check merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).set_checked(checked=True)

    with allure.step("Try to save"):
        settings.press_save()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(alert_merge)

    with allure.step("Uncheck merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).uncheck()

    with allure.step("Check diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).set_checked(checked=True)

    with allure.step("Try to save"):
        settings.press_save()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(alert_diarization)

    with allure.step("Uncheck diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).uncheck()

    with allure.step("Change parameters"):
        page.locator(RECOGNITION_PRIORITY).locator('[type="number"]').fill("10")
        page.locator(CHECKBOX_ECONOMIZE).set_checked(checked=True)
        #page.locator(CHECKBOX_USE_WEBHOOK).set_checked(checked=True)

    with allure.step("Press (Save)"):
        settings.press_save()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(data_updated)

    with allure.step("Page reload"):
        settings.reload_page()

    with allure.step("Check settings"):
        #expect(page.locator(SELECT_ENGINE)).to_contain_text("assembly_ai")
        #expect(page.locator(SELECT_MODEL)).to_contain_text("best")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("10")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        #expect(page.locator(CHECKBOX_USE_WEBHOOK)).to_be_checked()
        #expect(page.locator(CHECKBOX_ADD_PUNCTUATION)).to_be_checked()
        expect(page.locator(CHECKBOX_ENGINE_DIARIZATION)).to_be_checked()
        expect(page.locator(CHECKBOX_NORMALIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PROFANITY_FILTER)).not_to_be_checked()
        expect(page.locator(CHECKBOX_LITERATURE_STYLE)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PHONE_FORMATTING)).to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(8)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_check_re_recognize_in_actions_with_calls")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_check_re_recognize_in_actions_with_calls")
def test_check_re_recognize_in_actions_with_calls(base_url, page: Page) -> None:
    settings = Settings(page)

    # expected_languages = ("Английский (Великобритания)Английский (США)Испанский (Латинская Америка, Карибский регион, "
    #                       "код региона UN M49)Испанский (Испания)Французский (Франция)Португальский "
    #                       "(Бразилия)Португальский (Португалия)РусскийТурецкий (Турция)УкраинскийУзбекскийАвто")

    # expected_engines = "DeepgramGigaAMHappyscribeNLab SpeechIMOT.IOwhisperЯндексЯндекс v3"

    alert_merge = "Опция 'Объединить дорожки в один файл' не может быть выбрана одновременно с любой из диаризаций"

    alert_diarization = "Выберите только одну диаризацию среди опций"

    alert_fill_parameters = "Заполните необходимые параметры"

    action_started = "Работа пошла. Отслеживание ниже."

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to actions with calls"):
        settings.click_actions_with_calls()

    with allure.step("Click to actions select"):
        page.locator(BLOCK_ACTION_SELECT).get_by_text("Выберите действие").click(force=True)
        page.wait_for_selector(MENU)

    with allure.step("Choose re-recognize in action select"):
        settings.choose_option(2)
        page.wait_for_selector(SELECT_LANGUAGE)

#  check all combinations of engines and models

    with allure.step("Click to language"):
        settings.click_language_select()

    with allure.step("Check language list"):
        expect(page.locator(MENU)).to_contain_text(fucking_stupidity)

    with allure.step("Close language menu"):
        page.locator('[class*="STT_order_"]').click()

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Deepgram"):
        settings.choose_option(1)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Обобщённаяwhisper")

    with allure.step("Select model Обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model whisper"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Deepgram")
        expect(page.locator(SELECT_MODEL)).to_contain_text("whisper")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Happyscribe"):
        settings.choose_option(3)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Happyscribe")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose NLab Speech"):
        settings.choose_option(4)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("ОбобщённаяЖадный")

    with allure.step("Select model Обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Жадный"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("NLab Speech")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Жадный")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose IMOT.IO"):
        settings.choose_option(5)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("IMOT.IO")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose whisper"):
        settings.choose_option(6)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Стандарт")

    with allure.step("Select model Стандарт"):
        settings.choose_option(0)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("whisper")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Стандарт")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).not_to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(5)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose Яндекс"):
        settings.choose_option(7)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Обобщённая"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(3)
    #
    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose yandex_v3"):
        settings.choose_option(8)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Check model list"):
        expect(page.locator(MENU)).to_contain_text("Отложенная обобщённаяОбобщённая")

    with allure.step("Select model Отложенная обобщённая"):
        settings.choose_option(0)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model Обобщённая"):
        settings.choose_option(1)

    with allure.step("Check engine parameters"):
        expect(page.locator(SELECT_ENGINE)).to_contain_text("Яндекс v3")
        expect(page.locator(SELECT_MODEL)).to_contain_text("Обобщённая")
        expect(page.locator(CHECKBOX_MERGE_ALL_TO_ONE)).to_be_checked()
        expect(page.locator(RECOGNITION_PRIORITY).locator('[type="number"]')).to_have_value("1")
        expect(page.locator(CHECKBOX_DIARIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_ENGINE_DIARIZATION)).to_be_checked()
        expect(page.locator(CHECKBOX_NORMALIZATION)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PROFANITY_FILTER)).not_to_be_checked()
        expect(page.locator(CHECKBOX_LITERATURE_STYLE)).not_to_be_checked()
        expect(page.locator(CHECKBOX_PHONE_FORMATTING)).to_be_checked()
        expect(page.locator(CHECKBOX_ECONOMIZE)).not_to_be_checked()
        expect(page.locator('[type="checkbox"]')).to_have_count(8)

#  check save combinations
    with allure.step("Choose dates"):
        settings.choose_period_date(today.strftime("%d/%m/%Y"), today.strftime("%d/%m/%Y"))

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose assembly_ai"):
        settings.choose_option(7)

    with allure.step("Click to engine"):
        settings.click_engine_select()

    with allure.step("Choose assembly_ai"):
        settings.choose_option(8)

    with allure.step("Uncheck merge"):
        page.wait_for_timeout(500)
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).uncheck()

    with allure.step("Check (Save) button is enabled"):
        page.wait_for_timeout(500)
        expect(page.locator('[class="flex-end"]').locator('[type="button"]')).to_be_enabled()

    with allure.step("Try to save"):
        page.locator('[class="flex-end"]').locator('[type="button"]').click()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(alert_fill_parameters)

    with allure.step("Click to model"):
        settings.click_model_select()

    with allure.step("Select model best"):
        settings.choose_option(0)

    with allure.step("Check (Save) button is enabled"):
        expect(page.locator('[class="flex-end"]').locator('[type="button"]')).to_be_enabled()

    with allure.step("Check merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).set_checked(checked=True)

    with allure.step("Try to save"):
        page.locator('[class="flex-end"]').locator('[type="button"]').click()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(alert_merge)

    with allure.step("Uncheck merge all to one checkbox"):
        page.locator(CHECKBOX_MERGE_ALL_TO_ONE).uncheck()

    with allure.step("Check diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).set_checked(checked=True)

    with allure.step("Try to save"):
        page.locator('[class="flex-end"]').locator('[type="button"]').click()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(alert_diarization)

    with allure.step("Uncheck diarization checkbox"):
        page.locator(CHECKBOX_DIARIZATION).uncheck()

    with allure.step("Change parameters"):
        page.locator(RECOGNITION_PRIORITY).locator('[type="number"]').fill("10")
        page.locator(CHECKBOX_ECONOMIZE).set_checked(checked=True)
        #page.locator(CHECKBOX_USE_WEBHOOK).set_checked(checked=True)

    with allure.step("Press (Save)"):
        page.locator('[class="flex-end"]').locator('[type="button"]').click()

    with allure.step("Wait for alert and check alert message"):
        settings.check_alert(action_started)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_tariffication_if_empty")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_tariffication_history. check have warning if []")
def test_user_tariffication_if_empty(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_TARIFFICATION).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(MESSAGE_TARIFFICATION_EMPTY)).to_contain_text("Нет подключенных тарифов")

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_WRITEOFFS).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(MESSAGE_TARIFFICATION_EMPTY)).to_contain_text("Нет информации о списаниях")
        expect(page.locator('[placeholder="Поиск по тарифу или услуге"]')).to_have_count(1)
        expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CHARGES).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(MESSAGE_TARIFFICATION_EMPTY)).to_contain_text("Нет информации о платежах")
        expect(page.locator('[placeholder="Поиск по договору и назначению платежа"]')).to_have_count(1)
        expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_tariffication_if_500")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_tariffication_history. check have warning if 500")
def test_user_tariffication_if_500(base_url, page: Page) -> None:
    settings = Settings(page)

    error_first_line = "Возникла ошибка в процессе формирования таблицы"
    error_second_line = "Пожалуйста, перезагрузите страницу браузера или зайдите в систему заново"

    def handle_tariff(route: Route):
        route.fulfill(status=500, body="")
    # Intercept the route
    page.route("**/billing/billing_info", handle_tariff)
    page.route("**/billing/charges", handle_tariff)
    page.route("**/billing/payments", handle_tariff)

    # fix after https://task.imot.io/browse/DEV-3130

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to consumption history"):
        page.locator(BUTTON_TARIFFICATION).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(MESSAGE_TARIFFICATION_EMPTY)).to_contain_text("Возникла ошибка, пожалуйста, обновите страницу")

    with allure.step("Go to consumption history GPT"):
        page.locator(BUTTON_WRITEOFFS).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error_first_line)
        expect(page.locator(CONSUMPTION_ERROR_SECOND_LINE)).to_contain_text(error_second_line)
        #expect(page.locator('[placeholder="Поиск по тарифу или услуге"]')).to_have_count(1)
        #expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CHARGES).click()
        page.wait_for_timeout(500)
        page.wait_for_selector(MESSAGE_TARIFFICATION_EMPTY)

    with allure.step("Check snackbar"):
        settings.check_alert("Ошибка 500: Внутренняя ошибка сервера")

    with allure.step("Check exist search, calendar, mocked data and total count"):
        expect(page.locator(CONSUMPTION_ERROR_FIRST_LINE)).to_contain_text(error_first_line)
        expect(page.locator(CONSUMPTION_ERROR_SECOND_LINE)).to_contain_text(error_second_line)
        #expect(page.locator('[placeholder="Поиск по договору и назначению платежа"]')).to_have_count(1)
        #expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_user_tariffication")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_user_tariffication. mocked")
def test_user_tariffication(base_url, page: Page) -> None:
    settings = Settings(page)

    tariff_block_contain_text = ("О тарифахСписанияПлатежиИндивидуальный тарифТекущий баланс:0 рубОбщая сумма:60 000 руб"
                                 "Договор:0009Срок действия:БессрочноПериод тарификации:1 мес.Запросить счет"
                                 "Включено в тарифЧек-листы звонкаОповещенияАбонентская плата")

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    # mocks
    def handle_tariff(route: Route):
        json_tariff = [
            {
                "agreement": {
                    "agreement_id": 835,
                    "agreement_number": "0009",
                    "balance": 0.0,
                    "create_date": "2024-10-01",
                    "credit": 0.0,
                    "currency_symbol": "руб",
                    "organization_id": 486,
                    "pay_code": None,
                    "user_id": 602
                },
                "tariff": {
                    "archive": 0,
                    "assign_unavailable": 0,
                    "assigned": 1,
                    "balance_blocks": 1,
                    "block_to_close_days": 0,
                    "client_unavailable": 0,
                    "currency_id": 1,
                    "dynamic_rent": 0,
                    "float_period": 0,
                    "link": "https://docs.google.com/document/d/1HBTqEQw6IAcNJcJN50HqDSZyzNU8uGn3cGM0IguiOD4/edit",
                    "name": "АРХИВ. Пакет минут (расширенный)",
                    "outer_id": "minutes_advanced",
                    "payment_type": 0,
                    "tariff_id": 41
                },
                "subscriptions": [
                    {
                        "agreement_id": 835,
                        "agreement_number": "0009",
                        "currency_symbol": "руб",
                        "current_blocking": 0,
                        "description": "Индивидуальный тариф",
                        "err_details": "",
                        "last_tariffication_period_end": "2024-10-31",
                        "last_tariffication_period_start": "2024-10-10",
                        "next_change_date": None,
                        "period_id": 1,
                        "period_length": 1,
                        "period_type": 3,
                        "services": [
                            {
                                "bulk_service": False,
                                "hidden_for_user": 0,
                                "multiplicator": 1.0,
                                "outer_id": "check_list_call",
                                "periodic_service": 1,
                                "price": 10000.0,
                                "service_description": "Чек-листы звонка",
                                "service_id": 613,
                                "tariff_service_id": 2,
                                "weight": 0
                            },
                            {
                                "bulk_service": False,
                                "hidden_for_user": 0,
                                "multiplicator": 1.0,
                                "outer_id": "notifications",
                                "periodic_service": 1,
                                "price": 10000.0,
                                "service_description": "Оповещения",
                                "service_id": 614,
                                "tariff_service_id": 3,
                                "weight": 0
                            },
                            {
                                "bulk_service": False,
                                "hidden_for_user": 0,
                                "multiplicator": 1.0,
                                "outer_id": "subscription_fee",
                                "periodic_service": 1,
                                "price": 40000.0,
                                "service_description": "Абонентская плата",
                                "service_id": 615,
                                "tariff_service_id": 6,
                                "weight": 0
                            },
                            {
                                "bulk_service": True,
                                "hidden_for_user": 1,
                                "multiplicator": 1.0,
                                "outer_id": "quote_overspending",
                                "periodic_service": 1,
                                "price": 0.0,
                                "service_description": "Квота по перерасходу (счетчик)",
                                "service_id": 616,
                                "tariff_service_id": 9,
                                "weight": 0
                            },
                            {
                                "bulk_service": True,
                                "hidden_for_user": 1,
                                "multiplicator": 1.0,
                                "outer_id": "quote_consumption",
                                "periodic_service": 1,
                                "price": 0.0,
                                "service_description": "Квота по потреблению (счетчик)",
                                "service_id": 617,
                                "tariff_service_id": 10,
                                "weight": 0
                            }
                        ],
                        "state": 2,
                        "subscription_end_date": "9999-12-31 23:59:59",
                        "subscription_id": 239,
                        "subscription_instance_end_date": "9999-12-31 23:59:59",
                        "subscription_instance_start_date": "2024-10-10 00:00:00",
                        "subscription_start_date": "2024-10-10 00:00:00",
                        "tariff_descr": "АРХИВ. Пакет минут (расширенный)",
                        "tariff_id": 41,
                        "total_price": 60000.0
                    }
                ]
            }
        ]
        # fulfill the route with the mock data
        route.fulfill(json=json_tariff)
        # Intercept the route
    page.route("**/billing/billing_info", handle_tariff)

    def handle_writeoffs(route: Route):
        json_writeoffs = [
            {
                "agreement_id": 835,
                "agreement_number": "0009",
                "amount": 10000.0,
                "charge_date": "2024-10-10 11:46:25",
                "charge_id": 72940,
                "currency_id": 1,
                "currency_symbol": "руб",
                "is_bulk_service": False,
                "is_connection_charge": 0,
                "period": "2024-10-10",
                "periodic_service": 1,
                "service_id": 613,
                "subscription_description": "Индивидуальный тариф",
                "subscription_id": 239,
                "subscription_instance_id": 239,
                "tariff_description": "АРХИВ. Пакет минут (расширенный)",
                "tariff_id": 41,
                "tariff_service_description": "Чек-листы звонка",
                "tariff_service_id": 2,
                "tarification_period_end": "2024-10-31",
                "tarification_period_start": "2024-10-01",
                "usage_period": "2024-10-10",
                "user_name": "Клиент тестовый 2",
                "volume": 1.0
            },
            {
                "agreement_id": 835,
                "agreement_number": "0009",
                "amount": 10000.0,
                "charge_date": "2024-10-10 11:46:25",
                "charge_id": 72941,
                "currency_id": 1,
                "currency_symbol": "руб",
                "is_bulk_service": False,
                "is_connection_charge": 0,
                "period": "2024-10-10",
                "periodic_service": 1,
                "service_id": 614,
                "subscription_description": "Индивидуальный тариф",
                "subscription_id": 239,
                "subscription_instance_id": 239,
                "tariff_description": "АРХИВ. Пакет минут (расширенный)",
                "tariff_id": 41,
                "tariff_service_description": "Оповещения",
                "tariff_service_id": 3,
                "tarification_period_end": "2024-10-31",
                "tarification_period_start": "2024-10-01",
                "usage_period": "2024-10-10",
                "user_name": "Клиент тестовый 2",
                "volume": 1.0
            },
            {
                "agreement_id": 835,
                "agreement_number": "0009",
                "amount": 40000.0,
                "charge_date": "2024-10-10 11:46:25",
                "charge_id": 72942,
                "currency_id": 1,
                "currency_symbol": "руб",
                "is_bulk_service": False,
                "is_connection_charge": 0,
                "period": "2024-10-10",
                "periodic_service": 1,
                "service_id": 615,
                "subscription_description": "Индивидуальный тариф",
                "subscription_id": 239,
                "subscription_instance_id": 239,
                "tariff_description": "АРХИВ. Пакет минут (расширенный)",
                "tariff_id": 41,
                "tariff_service_description": "Абонентская плата",
                "tariff_service_id": 6,
                "tarification_period_end": "2024-10-31",
                "tarification_period_start": "2024-10-01",
                "usage_period": "2024-10-10",
                "user_name": "Клиент тестовый 2",
                "volume": 1.0
            }

        ]
        # fulfill the route with the mock data
        route.fulfill(json=json_writeoffs)
        # Intercept the route
    page.route("**/billing/charges", handle_writeoffs)

    def handle_charges(route: Route):
        json_charges = [
            {
                "agreement_id": 835,
                "agreement_number": "0009",
                "amount": 60000.0,
                "auto_payment_id": None,
                "cancel_date": None,
                "comment": "Some text about pyment",
                "currency_symbol": "руб",
                "document_id": None,
                "inn": None,
                "is_vps": 0,
                "local_date": "2024-10-10 11:46:12",
                "manager_fio": "Анастасия Соколова",
                "manager_id": 2,
                "order_number": None,
                "pay_code": None,
                "pay_date": "2024-10-01 11:46:12",
                "payment_id": 1413,
                "payment_order_number": None,
                "status": 0,
                "unqiue_payment_id": "20241046114612-9683",
                "user_id": 602,
                "user_login": "login-X4UYT",
                "user_name": "Клиент тестовый 2"
            }
        ]
        # fulfill the route with the mock data
        route.fulfill(json=json_charges)
        # Intercept the route
    page.route("**/billing/payments", handle_charges)


    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with user"):
        settings.auth(LOGIN_USER, PASSWORD)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to tarriffication history"):
        page.locator(BUTTON_TARIFFICATION).click()

    with allure.step("Check tariff name"):
        expect(page.locator('[class*="styles_body"]')).to_contain_text(tariff_block_contain_text)

    with allure.step("Check invoice request"):
        page.get_by_role("button", name="Запросить счет").click()
        page.wait_for_selector(MODAL_WINDOW)
        expect(page.locator(MODAL_WINDOW).get_by_role("button", name="Запросить")).to_have_count(1)
        page.locator(MODAL_WINDOW).get_by_role("button", name="Отмена").click()

    with allure.step("Go to writeoffs"):
        page.locator(BUTTON_WRITEOFFS).click()

    with allure.step("Check exist search for chats, calendar, mocked data and total count"):
        expect(page.locator('[placeholder="Поиск по тарифу или услуге"]')).to_have_count(1)
        expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)
        #  check first row
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="1"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="2"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="3"]')).to_contain_text("АРХИВ. Пакет минут (расширенный)")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_contain_text("Чек-листы звонка")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="5"]')).to_contain_text("1")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="6"]')).to_contain_text("10 000")

        #  check second row
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="1"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="2"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="3"]')).to_contain_text("АРХИВ. Пакет минут (расширенный)")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="4"]')).to_contain_text("Оповещения")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="5"]')).to_contain_text("1")
        expect(page.locator('[aria-rowindex="3"]').locator('[aria-colindex="6"]')).to_contain_text("10 000")

        #  check third row
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="1"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="2"]')).to_contain_text("10.10.2024")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="3"]')).to_contain_text("АРХИВ. Пакет минут (расширенный)")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="4"]')).to_contain_text("Абонентская плата")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="5"]')).to_contain_text("1")
        expect(page.locator('[aria-rowindex="4"]').locator('[aria-colindex="6"]')).to_contain_text("40 000")

        # #  check total count
        expect(page.locator(TOTAL_IN_TABLE)).to_contain_text("60 000")

    with allure.step("Fill search by ya"):
        page.locator(SEARCH_IN_TARIFFICATION).locator('[type="text"]').fill("або")

    with allure.step("Check search that chat_gpt not visible"):
        expect(page.locator('[class*="style_table_"]')).not_to_contain_text("10 000")

    with allure.step("Go to consumption history chats"):
        page.locator(BUTTON_CHARGES).click()

    with allure.step("Check exist search, calendar, mocked data and total count"):

        expect(page.locator('[placeholder="Поиск по договору и назначению платежа"]')).to_have_count(1)
        expect(page.locator(SEARCH_IN_TARIFFICATION)).to_have_count(1)
        #  check first row
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="1"]')).to_contain_text("01.10.2024")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="2"]')).to_contain_text("0009")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="3"]')).to_contain_text("Some text about pyment")
        expect(page.locator('[aria-rowindex="2"]').locator('[aria-colindex="4"]')).to_contain_text("+60 000")
        #  check total count
        expect(page.locator(TOTAL_IN_TABLE)).to_contain_text("60 000")

    # with allure.step("Fill search by chat1"):
    #     page.locator(SEARCH_IN_CONSUMPTION_CHATS).locator('[type="text"]').fill("chat1")
    #
    # with allure.step("Check search that chat_gpt not visible"):
    #     expect(page.locator('[class*="communicationsStyles_table_"]')).not_to_contain_text("chat2")

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_change_role_for_user_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_change_role_for_user_by_admin")
def test_change_role_for_user_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Change role for manager"):
        settings.change_role("Интегратор")

    with allure.step("Press save"):
        settings.press_save()

    with allure.step("Reload page"):
        settings.reload_page()

    with allure.step("Check that role changed"):
        expect(page.locator(SELECT_ROLE)).to_contain_text("Интегратор")

    with allure.step("Change role for manager"):
        settings.change_role("Администратор")

    with allure.step("Press save"):
        settings.press_save()

    with allure.step("Reload page"):
        settings.reload_page()

    with allure.step("Check that role changed"):
        expect(page.locator(SELECT_ROLE)).to_contain_text("Администратор")

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)

#############
# @pytest.mark.e2e
# @pytest.mark.settings
# @allure.title("test_access_right_restt_for_user")
# @allure.severity(allure.severity_level.CRITICAL)
# @allure.description("test_change_role_for_user_by_admin")
# def test_access_right_restt_for_user(base_url, page: Page) -> None:
#     settings = Settings(page)
#
#     with allure.step("Create admin"):
#         USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)
#
#     with allure.step("Create user"):
#         USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)
#
#     with allure.step("Go to page"):
#         settings.navigate(base_url)
#
#     with allure.step("Auth with admin"):
#         settings.auth(LOGIN_USER, PASSWORD)
#
#     with allure.step("Expand call"):
#
#
#     with allure.step("Chek that no any stt buttons in calls"):
#         expect(page.locator('[data-testid="calls_actions_retag"]')).to_have_count(0)
#         expect(page.locator('[aria-label="Перетегировать"]')).to_have_count(0)
#
#     with allure.step("Change access_right"):
#         rights = {"restt":True,"delete_call":False,"call_info_processing":False}
#         give_access_right(API_URL, TOKEN_ADMIN, USER_ID_USER, rights)
#
#     with allure.step("Reload page"):
#         settings.reload_page()
#
#     with allure.step("Chek that no any stt buttons in calls"):
#         expect(page.locator('[data-testid="calls_actions_retag"]')).to_have_count(1)
#         expect(page.locator('[aria-label="Перетегировать"]')).to_have_count(1)
#
#
#     with allure.step("Delete admin"):
#         delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)
#
#     with allure.step("Delete user"):
#         delete_user(API_URL, TOKEN_USER, USER_ID_USER)


@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_upload_file_for_user_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_upload_file_for_user_by_admin")
def test_upload_file_for_user_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to upload"):
        settings.click_to_upload_files()

    with allure.step("Check requirements text"):
        expect(page.locator(BUTTON_CREATE_COMMUNICATIONS)).to_be_disabled()
        expect(page.locator('[class*="_requirements_"]')).to_have_text("Требования к файламФормат: wav, mp3, opus, ogg, flac, aiffРазмер одного файла: до 1 гб")

    with allure.step("Upload file"):
        page.locator('[name="audio"]').set_input_files("audio/stereo.opus")
        page.wait_for_timeout(6000)

    with allure.step("Check"):
        expect(page.locator(BUTTON_CREATE_COMMUNICATIONS)).to_be_enabled()
        expect(page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS)).to_be_enabled()
        expect(page.locator('[title="stereo.opus"]')).to_have_count(1)

    with allure.step("Use (Delete all) button"):
        page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS).click()
        page.wait_for_selector(MODAL_WINDOW)

    with allure.step("Accept in modal window"):
        page.locator(MODAL_WINDOW).locator('[data-testid="upload_delete_all_ok"]').click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Check alert"):
        settings.check_alert("Файлы успешно удалены")

    with allure.step("Check"):
        expect(page.locator(BUTTON_CREATE_COMMUNICATIONS)).to_be_disabled()
        expect(page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS)).not_to_be_visible()
        expect(page.locator('[title="stereo.opus"]')).to_have_count(0)

    with allure.step("Upload file"):
        page.locator('[name="audio"]').set_input_files("audio/stereo.opus")
        page.wait_for_timeout(8000)

    with allure.step("Delete from list"):
        page.locator('[title="Remove file"]').click()
        page.wait_for_selector(MODAL_WINDOW)

    with allure.step("Accept in modal window"):
        page.locator(MODAL_WINDOW).locator('[data-testid="upload_delete_ok"]').click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Check alert"):
        settings.check_alert("Файл успешно удален")

    with allure.step("Check"):
        expect(page.locator(BUTTON_CREATE_COMMUNICATIONS)).to_be_disabled()
        expect(page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS)).not_to_be_visible()
        expect(page.locator('[title="stereo.opus"]')).to_have_count(0)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)



@pytest.mark.e2e
@pytest.mark.settings
@allure.title("test_upload_errors_file_for_user_by_admin")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_upload_errors_file_for_user_by_admin")
def test_upload_errors_file_for_user_by_admin(base_url, page: Page) -> None:
    settings = Settings(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD)

    with allure.step("Go to page"):
        settings.navigate(base_url)

    with allure.step("Auth with admin"):
        settings.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        settings.go_to_user(LOGIN_USER)

    with allure.step("Go to settings"):
        settings.click_settings()

    with allure.step("Go to upload"):
        settings.click_to_upload_files()

    with allure.step("Upload file"):
        page.locator('[name="audio"]').set_input_files("audio/2G.opus")
        page.wait_for_timeout(2000)

    with allure.step("Check alert for file more than 1G"):
        settings.check_alert("Размер загружаемого файла превышает  1 гб")

    with allure.step("Upload file with wrong format"):
        page.locator('[name="audio"]').set_input_files("audio/text.txt")
        page.wait_for_timeout(2000)

    with allure.step("Check alert for file with wrong format"):
        settings.check_alert("Формат файла не соответствует требованиям")

    with allure.step("Upload normal audio"):
        page.locator('[name="audio"]').set_input_files("audio/stereo.opus")
        page.wait_for_timeout(2000)

    with allure.step("Check that normal audio downloaded"):
        expect(page.locator('[title="stereo.opus"]')).to_have_count(1)

    with allure.step("Upload normal audio again"):
        page.locator('[name="audio"]').set_input_files("audio/stereo.opus")
        page.wait_for_timeout(2000)

    with allure.step("Check alert for file which was already uploaded"):
        settings.check_alert("Загружен дубликат файла")

# clear
    with allure.step("Use (Delete all) button"):
        page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS).click()
        page.wait_for_selector(MODAL_WINDOW)

    with allure.step("Accept in modal window"):
        page.locator(MODAL_WINDOW).locator('[data-testid="upload_delete_all_ok"]').click()
        page.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Check alert"):
        settings.check_alert("Файлы успешно удалены")

    with allure.step("Check"):
        expect(page.locator(BUTTON_CREATE_COMMUNICATIONS)).to_be_disabled()
        expect(page.locator(BUTTON_DELETE_ALL_COMMUNICATIONS)).not_to_be_visible()
        expect(page.locator('[title="2G.opus"]')).to_have_count(0)
        expect(page.locator('[title="text.txt"]')).to_have_count(0)
        expect(page.locator('[title="stereo.opus"]')).to_have_count(0)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)