from utils.variables import *
from pages.deals import *
import pytest
from utils.dates import today
from utils.create_delete_user import create_user, delete_user
import allure
import os
from openpyxl import load_workbook


@pytest.mark.e2e
@pytest.mark.deals
@allure.title("test_deal")
@allure.severity(allure.severity_level.CRITICAL)
@allure.description("test_deal")
def test_deal(base_url, page: Page, context: BrowserContext) -> None:
    deals = Deals(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True, deal_check_list=True)

    with allure.step("Go to url"):
        deals.navigate("http://192.168.10.101/feature-dev-3474/")

    with allure.step("Auth with admin"):
        deals.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        deals.go_to_user(LOGIN_USER)

    with allure.step("Go to deals"):
        deals.click_deals()

    with allure.step("Check deal in list"):
        expect(deals.deals_found).to_have_text("Найдено сделок 1 из 1")
        expect(page.get_by_text("0987654321")).to_have_count(1)
        expect(deals.deal_date).to_contain_text(today.strftime("%d.%m.%Y"))
        expect(deals.communications_count).to_have_text("1")
        expect(deals.score_percent).to_have_text("90%")
        expect(deals.deal_score).to_have_text("90 баллов")

    with allure.step("Open new tab"):
        with context.expect_page() as new_tab_event:
            deals.button_share_call.click()
            open_deal=new_tab_event.value

    with allure.step("Check"):
        open_deal.wait_for_timeout(2000)
        open_deal.wait_for_selector('[alt="Imot.io loader"]', state="hidden")
        open_deal.wait_for_timeout(2000)
        opened_deal = Deals(open_deal)

        # deal communication
        expect(opened_deal.deal_communication_date).to_contain_text(today.strftime("%d.%m.%y"))
        expect(opened_deal.deal_communication_score_percent).to_have_text("0%")
        expect(opened_deal.deal_communication_score).to_have_text("0 баллов")
        expect(opened_deal.deal_communication_operator_phone).to_have_text("1234567890")
        expect(opened_deal.deal_communication_client_phone).to_have_text("0987654321")
        expect(opened_deal.deal_communication_duration).to_have_text("00:00:38")
        expect(opened_deal.button_share_call).to_have_count(1)
        expect(opened_deal.button_expand_call).to_have_count(1)
        expect(opened_deal.blue_tag).to_have_count(3)

        # deal buttons
        expect(opened_deal.deal_button_retag).to_have_count(1)
        expect(opened_deal.button_calls_list_download).to_have_count(1)

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)



@pytest.mark.deals
@pytest.mark.e2e
@allure.title("test_check_download_button_in_open_deal")
@allure.severity(allure.severity_level.NORMAL)
@allure.description("test_check_download_button_in_open_deal")
def test_check_download_button_in_open_deal(base_url, page: Page, context: BrowserContext) -> None:
    deals = Deals(page)

    with allure.step("Create admin"):
        USER_ID_ADMIN, TOKEN_ADMIN, LOGIN_ADMIN = create_user(API_URL, ROLE_ADMIN, PASSWORD)

    with allure.step("Create user"):
        USER_ID_USER, TOKEN_USER, LOGIN_USER = create_user(API_URL, ROLE_USER, PASSWORD, upload_call=True, deal_check_list=True)

    with allure.step("Go to url"):
        deals.navigate("http://192.168.10.101/feature-dev-3474/")

    with allure.step("Auth with admin"):
        deals.auth(LOGIN_ADMIN, PASSWORD)

    with allure.step("Go to user"):
        deals.go_to_user(LOGIN_USER)

    with allure.step("Go to deals"):
        deals.click_deals()

    with allure.step("Open deal"):
        with context.expect_page() as new_tab_event:
            deals.button_share_call.click()
            open_deal=new_tab_event.value

    with allure.step("Check"):
        open_deal.wait_for_timeout(2000)
        open_deal.wait_for_selector('[alt="Imot.io loader"]', state="hidden")
        open_deal.wait_for_timeout(2000)
        opened_deal = Deals(open_deal)

    with allure.step("Press button (Download)"):
        opened_deal.download_deal()

    with allure.step("Check content of button download"):
        expect(opened_deal.menu).to_have_text("Экспорт аудиоЭкспорт аудио (многоканальное)Экспорт расшифровкиЭкспорт коммуникаций")

    with allure.step("Choose (Export audio) option from opened menu"):
        # Start waiting for the download
        with open_deal.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            opened_deal.menu.get_by_text("Экспорт аудио", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert download.suggested_filename == "deal_audio_mono.zip"
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 120000 < os.path.getsize(path + download.suggested_filename) < 160000

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Press button (Download)"):
        opened_deal.download_deal()

    with allure.step("Choose (Export audio) option from opened menu"):
        # Start waiting for the download
        with open_deal.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            opened_deal.menu.get_by_text("Экспорт аудио (многоканальное)", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert download.suggested_filename == "deal_audio_stereo.zip"
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 100 < os.path.getsize(path + download.suggested_filename) < 10000

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Press button (Download)"):
        opened_deal.download_deal()

    with allure.step("Press (Export)"):
        opened_deal.menu.get_by_text("Экспорт расшифровки", exact=True).click()
        open_deal.wait_for_selector(MODAL_WINDOW)

    with allure.step("Choose (Export transcribe) option from opened menu"):
        # Start waiting for the download
        with open_deal.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            opened_deal.modal_window.get_by_text("Экспортировать", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'
        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert download.suggested_filename == "deal_calls_text.xlsx"
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 5500 < os.path.getsize(path + download.suggested_filename) < 8000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Экспорт расшифровки звонков"
        assert sheet.max_row == 33
        assert sheet.max_column == 3

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Close modal with export"):
        open_deal.wait_for_selector(BUTTON_CROSS, timeout=wait_until_visible)
        open_deal.locator(BUTTON_CROSS).click()
        open_deal.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Press button (Download)"):
        opened_deal.download_deal()

    with allure.step("Press (Export)"):
        opened_deal.menu.get_by_text("Экспорт коммуникаций", exact=True).click()
        open_deal.wait_for_selector(MODAL_WINDOW)

    with allure.step("Choose (Export transcribe) option from opened menu"):
        # Start waiting for the download
        with open_deal.expect_download(timeout=60000) as download_info:
            # Perform the action that initiates download
            opened_deal.modal_window.get_by_text("Экспортировать", exact=True).click()
        download = download_info.value
        path = f'{os.getcwd()}/'

        # Wait for the download process to complete and save the downloaded file somewhere
        download.save_as(path + download.suggested_filename)

    with allure.step("Check that export (zip) downloaded"):
        assert download.suggested_filename == "deal_audio_calls_list.xlsx"
        assert os.path.isfile(path + download.suggested_filename) == True
        assert 5000 < os.path.getsize(path + download.suggested_filename) < 6000

    with allure.step("Check what we have inside excel"):
        wb = load_workbook(path + download.suggested_filename)
        sheet = wb.active

        assert sheet["A1"].value == "Выгрузка списка звонков"
        assert sheet.max_row == 3
        assert sheet.max_column == 14

    with allure.step("Remove downloaded export (zip)"):
        os.remove(path + download.suggested_filename)

    with allure.step("Check that downloaded export (zip) removed"):
        assert os.path.isfile(path + download.suggested_filename) == False

    with allure.step("Close modal with export"):
        open_deal.wait_for_selector(BUTTON_CROSS, timeout=wait_until_visible)
        open_deal.locator(BUTTON_CROSS).click()
        open_deal.wait_for_selector(MODAL_WINDOW, state="hidden")

    with allure.step("Delete admin"):
        delete_user(API_URL, TOKEN_ADMIN, USER_ID_ADMIN)

    with allure.step("Delete user"):
        delete_user(API_URL, TOKEN_USER, USER_ID_USER)